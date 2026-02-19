"""
ChannelPRO™ — Partner Revenue Optimizer (Multi-Tenant v4)
========================================================
Login → Client Intake → Step 1–4
Single instance, per-client data isolation, admin overview.

Architecture: the heavy lifting (auth, data I/O, scoring, AI) lives in
the ``utils/`` package.  This file wires up Streamlit page config,
login, sidebar navigation, and page rendering.
"""
import csv, io, json, os, re
import streamlit as st
import pandas as pd

try:
    from st_aggrid import AgGrid, GridOptionsBuilder, JsCode, GridUpdateMode, ColumnsAutoSizeMode
    HAS_AGGRID = True
except ImportError:
    HAS_AGGRID = False

# ── Module imports ──────────────────────────────────────────────────────
from utils.paths import (
    BASE_DIR, USERS_FILE, TENANTS_DIR,
    tenant_dir as _tenant_dir,
    all_tenants as _all_tenants,
    current_data_dir as _current_data_dir,
    save_path as _save_path,
    client_path as _client_path,
    csv_path as _csv_path,
    class_path as _class_path,
    raw_path as _raw_path,
    tenant_config_path as _tenant_config_path,
    be_path as _be_path,
    sd_path as _sd_path,
)
from utils.auth import (
    hash_pw as _hash_pw,
    verify_pw as _verify_pw,
    load_users as _load_users,
    save_users as _save_users,
    handle_login,
)
from utils.api import resolve_api_key, call_ai
from utils.data import (
    load_partners as _load_partners,
    invalidate_partner_cache,
    load_raw as _load_raw,
    save_raw as _save_raw,
    append_partner as _append_partner_raw,
    delete_partner as _delete_partner,
    partner_exists as _partner_exists,
    upsert_partner as _upsert_partner_raw,
    load_tenant_config as _load_tenant_config,
    save_tenant_config as _save_tenant_config,
    max_partners as _max_partners,
    partner_count as _partner_count,
    load_q_config as _load_q_config,
    save_q_config as _save_q_config,
    load_be as _load_be,
    save_be as _save_be,
)
from utils.scoring import (
    SCORECARD_METRICS, CATEGORIES, METRICS_BY_KEY, SC,
    METRIC_ALIASES, DEFAULT_Q_CONFIG, Q_LABELS, Q_DESCS,
    BE_SECTIONS, BE_SECTION_ICONS,
    calc_score, enabled as _enabled, grade as _grade, tiers as _tiers,
    synthetic_raw_for_score as _synthetic_raw_for_score,
    init_criteria as _init_criteria,
    ensure_criteria_complete as _ensure_criteria_complete,
    save_criteria as _save_criteria,
    rescore_all as _rescore_all,
    recalculate_benchmarks as _recalculate_benchmarks,
    calculate_revenue_recovery as _calculate_revenue_recovery,
    classify_partners,
    _sf,
)
from utils.ui import (
    YORK_LOGO_B64, LOGIN_BG_B64,
    inject_css,
    logo as _logo,
    brand as _brand,
    display_styled_assessment_table,
)

# ═════════════════════════════════════════════════════════════════════════
# COUNTRIES LIST
# ═════════════════════════════════════════════════════════════════════════
COUNTRIES = ["","Afghanistan","Albania","Algeria","Andorra","Angola","Antigua and Barbuda","Argentina","Armenia","Australia","Austria","Azerbaijan","Bahamas","Bahrain","Bangladesh","Barbados","Belarus","Belgium","Belize","Benin","Bhutan","Bolivia","Bosnia and Herzegovina","Botswana","Brazil","Brunei","Bulgaria","Burkina Faso","Burundi","Cabo Verde","Cambodia","Cameroon","Canada","Central African Republic","Chad","Chile","China","Colombia","Comoros","Congo","Costa Rica","Croatia","Cuba","Cyprus","Czech Republic","Denmark","Djibouti","Dominica","Dominican Republic","Ecuador","Egypt","El Salvador","Equatorial Guinea","Eritrea","Estonia","Eswatini","Ethiopia","Fiji","Finland","France","Gabon","Gambia","Georgia","Germany","Ghana","Greece","Grenada","Guatemala","Guinea","Guinea-Bissau","Guyana","Haiti","Honduras","Hungary","Iceland","India","Indonesia","Iran","Iraq","Ireland","Israel","Italy","Ivory Coast","Jamaica","Japan","Jordan","Kazakhstan","Kenya","Kiribati","Kosovo","Kuwait","Kyrgyzstan","Laos","Latvia","Lebanon","Lesotho","Liberia","Libya","Liechtenstein","Lithuania","Luxembourg","Madagascar","Malawi","Malaysia","Maldives","Mali","Malta","Marshall Islands","Mauritania","Mauritius","Mexico","Micronesia","Moldova","Monaco","Mongolia","Montenegro","Morocco","Mozambique","Myanmar","Namibia","Nauru","Nepal","Netherlands","New Zealand","Nicaragua","Niger","Nigeria","North Korea","North Macedonia","Norway","Oman","Pakistan","Palau","Palestine","Panama","Papua New Guinea","Paraguay","Peru","Philippines","Poland","Portugal","Qatar","Romania","Russia","Rwanda","Saint Kitts and Nevis","Saint Lucia","Saint Vincent and the Grenadines","Samoa","San Marino","Sao Tome and Principe","Saudi Arabia","Senegal","Serbia","Seychelles","Sierra Leone","Singapore","Slovakia","Slovenia","Solomon Islands","Somalia","South Africa","South Korea","South Sudan","Spain","Sri Lanka","Sudan","Suriname","Sweden","Switzerland","Syria","Taiwan","Tajikistan","Tanzania","Thailand","Timor-Leste","Togo","Tonga","Trinidad and Tobago","Tunisia","Turkey","Turkmenistan","Tuvalu","Uganda","Ukraine","United Arab Emirates","United Kingdom","United States","Uruguay","Uzbekistan","Vanuatu","Vatican City","Venezuela","Vietnam","Yemen","Zambia","Zimbabwe"]


# ═════════════════════════════════════════════════════════════════════════
# THIN WRAPPERS — keep old call-sites working with minimal changes
# ═════════════════════════════════════════════════════════════════════════
def _append_partner(row_dict, raw_dict):
    """Delegate to data module, passing currently enabled metrics."""
    _append_partner_raw(row_dict, raw_dict, _enabled())

def _upsert_partner(row_dict, raw_dict):
    """Delegate to data module, passing currently enabled metrics."""
    _upsert_partner_raw(row_dict, raw_dict, _enabled())

# ═════════════════════════════════════════════════════════════════════════
# AI AGENT HELPERS — Ask ChannelPRO  (stays here; tightly coupled to UI)
# ═════════════════════════════════════════════════════════════════════════
def _build_ai_system_prompt():
    """Build a system prompt containing all partner data and criteria definitions."""
    cr = st.session_state.get("criteria", {})
    em = _enabled(cr)
    partners = _load_partners()
    raw_all = _load_raw()
    raw_by_name = {r.get("partner_name",""): r for r in raw_all}

    criteria_lines = []
    for m in em:
        mc = cr.get(m["key"], {})
        if m["type"] == "quantitative":
            ranges = []
            for s in ("1","2","3","4","5"):
                r = mc.get("ranges",{}).get(s,{})
                lo, hi = r.get("min",""), r.get("max","")
                if lo and hi: ranges.append(f"  {s}: {lo}\u2013{hi}")
                elif lo: ranges.append(f"  {s}: \u2265{lo}")
                elif hi: ranges.append(f"  {s}: \u2264{hi}")
            criteria_lines.append(f"- {m['name']} (key: {m['key']}, unit: {m.get('unit','')}, {m['direction']}, quantitative)\n" + "\n".join(ranges))
        else:
            descs = [f"  {s}: {mc.get('descriptors',{}).get(s,'')}" for s in ("1","2","3","4","5")]
            criteria_lines.append(f"- {m['name']} (key: {m['key']}, {m['direction']}, qualitative)\n" + "\n".join(descs))

    partner_lines = []
    for p in partners:
        raw = raw_by_name.get(p.get("partner_name",""), {})
        metrics_str = []
        for m in em:
            mk = m["key"]
            try: score = int(p.get(mk,"") or 0)
            except: score = 0
            raw_val = raw.get(f"raw_{mk}", "")
            if score and raw_val:
                metrics_str.append(f"{m['name']}={score}/5 (raw:{raw_val})")
            elif score:
                metrics_str.append(f"{m['name']}={score}/5")
            else:
                metrics_str.append(f"{m['name']}=unscored")
        try: pct = float(p.get("percentage",0) or 0)
        except: pct = 0
        try: total = int(p.get("total_score",0) or 0)
        except: total = 0
        partner_lines.append(
            f"Partner: {p.get('partner_name','')}\n"
            f"  Tier: {p.get('partner_tier','N/A')} | Country: {p.get('partner_country','N/A')} | "
            f"City: {p.get('partner_city','N/A')} | PAM: {p.get('pam_name','N/A')} | "
            f"Discount: {p.get('partner_discount','N/A')}\n"
            f"  Total: {total} | Percentage: {pct:.1f}%\n"
            f"  Scores: {' | '.join(metrics_str)}"
        )

    system = f"""You are ChannelPRO\u2122 AI Assistant, an expert in partner channel management and analysis.
You analyze partner scorecard data and provide actionable insights.

SCORING SYSTEM: Each metric is scored 1-5 (5=best). Higher percentage = better overall performance.
Grade scale: A (\u226590%), B+ (\u226580%), B (\u226570%), C+ (\u226560%), C (\u226550%), D (<50%).

SCORING CRITERIA ({len(em)} active metrics):
{chr(10).join(criteria_lines)}

PARTNER DATA ({len(partners)} partners):
{chr(10).join(partner_lines) if partner_lines else "No partners scored yet."}

INSTRUCTIONS:
- Answer questions about partner performance, comparisons, filtering, and trends.
- When listing partners, include their key metrics and scores.
- You can suggest or make score updates when asked.
- Be specific with numbers and partner names.

Always respond with valid JSON (no markdown fences, no extra text) in this exact format:
{{
  "answer": "Your detailed analysis in plain text. Use \\n for line breaks.",
  "table": [
    {{"Partner": "Name", "Tier": "Gold", "Country": "US", "PAM": "Jane", "Total": 85, "Pct": "72.3%", "Key Metric": "value"}}
  ],
  "chart": {{
    "type": "bar or pie or hbar",
    "title": "Chart title",
    "x_label": "X axis label",
    "y_label": "Y axis label",
    "data": [{{"label": "Name", "value": 42.5}}, {{"label": "Name2", "value": 38.1}}]
  }},
  "updates": [
    {{"partner": "Partner Name", "metric_key": "metric_key_here", "new_score": 3, "reason": "Explanation"}}
  ]
}}

RULES for the JSON response:
- "answer" is ALWAYS required.
- "table" should be included when listing/filtering partners. Use null if not relevant.
- "chart" should be included when a visualization would help. Use null if not needed. Keep data to \u226415 items.
- "updates" should ONLY be included when the user explicitly asks to change/update scores. Use null otherwise.
- For "table", dynamically choose columns that are relevant to the query.
- For "chart", choose the best chart type: "bar" for comparisons, "pie" for distributions, "hbar" for ranked lists.
"""
    return system

def _call_ai(messages, api_key):
    """Call Anthropic API with conversation history (delegates to utils.api)."""
    system = _build_ai_system_prompt()
    return call_ai(messages, api_key, system)

def _render_ai_chart(chart_spec):
    """Render a chart from AI-generated spec using native Streamlit charts."""
    if not chart_spec or not chart_spec.get("data"): return
    title = chart_spec.get("title", "")
    data = chart_spec["data"]
    df = pd.DataFrame(data)
    if "label" not in df.columns or "value" not in df.columns: return
    if title:
        st.markdown(f"**{title}**")
    chart_df = df.set_index("label")
    try:
        st.bar_chart(chart_df)
    except Exception:
        st.dataframe(df, use_container_width=True, hide_index=True)

def _apply_ai_updates(updates, cr):
    """Apply score updates from AI response."""
    em = _enabled(cr)
    em_keys = {m["key"] for m in em}
    partners = _load_partners()
    raw_all = _load_raw()
    applied = 0
    for upd in updates:
        pn = upd.get("partner","")
        mk = upd.get("metric_key","")
        new_score = upd.get("new_score")
        if not pn or not mk or mk not in em_keys: continue
        if not isinstance(new_score, int) or new_score < 1 or new_score > 5: continue
        csv_p = next((p for p in partners if p.get("partner_name","").strip().lower() == pn.strip().lower()), None)
        raw_p = next((r for r in raw_all if r.get("partner_name","").strip().lower() == pn.strip().lower()), None)
        if not csv_p: continue
        csv_p[mk] = new_score
        if raw_p:
            raw_p[f"raw_{mk}"] = _synthetic_raw_for_score(mk, new_score, cr)
        si = {}
        for m in em:
            try: v = int(csv_p.get(m["key"],"") or 0)
            except: v = 0
            if v and 1 <= v <= 5: si[m["key"]] = v
        total = sum(si.values()); sn = len(si); mp = sn * 5
        csv_p["total_score"] = total; csv_p["max_possible"] = mp
        csv_p["percentage"] = round(total / mp * 100, 1) if mp else 0
        applied += 1
    if applied > 0:
        cp = _csv_path()
        if partners:
            fnames = list(partners[0].keys())
            with open(cp, "w", newline="") as f:
                w = csv.DictWriter(f, fieldnames=fnames, extrasaction="ignore")
                w.writeheader()
                for p in partners: w.writerow(p)
        rp = _raw_path()
        rp.write_text(json.dumps(raw_all, indent=2))
        invalidate_partner_cache()
    return applied


def _gen_xlsx(partners, enabled_metrics):
    """Generate an Excel workbook with the partner assessment heatmap."""
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    except ImportError: return None
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Partner Assessment"
    fills={1:PatternFill(start_color="FA7A7A",end_color="FA7A7A",fill_type="solid"),2:PatternFill(start_color="FA7A7A",end_color="FA7A7A",fill_type="solid"),3:PatternFill(start_color="FFFFCC",end_color="FFFFCC",fill_type="solid"),4:PatternFill(start_color="C6EFCE",end_color="C6EFCE",fill_type="solid"),5:PatternFill(start_color="C6EFCE",end_color="C6EFCE",fill_type="solid")}
    hf=PatternFill(start_color="1E2A3A",end_color="1E2A3A",fill_type="solid"); hfont=Font(color="FFFFFF",bold=True,size=10)
    bdr=Border(left=Side(style="thin",color="CCCCCC"),right=Side(style="thin",color="CCCCCC"),top=Side(style="thin",color="CCCCCC"),bottom=Side(style="thin",color="CCCCCC"))
    headers=["Rank","Partner Name","Tier","PAM","City","Country"]+[m["name"] for m in enabled_metrics]+["Total","%"]
    metric_start = 7
    metric_end = 6 + len(enabled_metrics)
    for c,h in enumerate(headers,1):
        cell=ws.cell(1,c,h); cell.fill=hf; cell.font=hfont; cell.border=bdr
        if metric_start <= c <= metric_end:
            cell.alignment=Alignment(horizontal="center",vertical="bottom",text_rotation=55,wrap_text=False)
            ws.column_dimensions[cell.column_letter].width=6
        else:
            cell.alignment=Alignment(horizontal="center",wrap_text=True)
            ws.column_dimensions[cell.column_letter].width=14 if c>6 else 18
    ws.row_dimensions[1].height=110
    ps=sorted(partners,key=lambda p:-int(p.get("total_score",0) or 0))
    for ri,p in enumerate(ps,2):
        ws.cell(ri,1,ri-1).border=bdr; ws.cell(ri,1).alignment=Alignment(horizontal="center")
        ws.cell(ri,2,p.get("partner_name","")).border=bdr
        ws.cell(ri,3,p.get("partner_tier","")).border=bdr
        ws.cell(ri,4,p.get("pam_name","")).border=bdr
        ws.cell(ri,5,p.get("partner_city","")).border=bdr
        ws.cell(ri,6,p.get("partner_country","")).border=bdr
        for ci,m in enumerate(enabled_metrics,7):
            try: v=int(p.get(m["key"],"") or 0)
            except: v=None
            cell=ws.cell(ri,ci); cell.border=bdr; cell.alignment=Alignment(horizontal="center")
            if v and 1<=v<=5: cell.value=v; cell.fill=fills[v]; cell.font=Font(bold=True)
        tc=len(enabled_metrics)+7
        try: tv=int(p.get("total_score",0) or 0)
        except: tv=0
        ws.cell(ri,tc,tv).border=bdr; ws.cell(ri,tc).font=Font(bold=True); ws.cell(ri,tc).alignment=Alignment(horizontal="center")
        try: pv=float(p.get("percentage",0) or 0)
        except: pv=0
        pc=ws.cell(ri,tc+1); pc.value=pv/100; pc.number_format="0.0%"; pc.border=bdr; pc.alignment=Alignment(horizontal="center"); pc.font=Font(bold=True)
    buf=io.BytesIO(); wb.save(buf); return buf.getvalue()


# ═════════════════════════════════════════════════════════════════════════
# PAGE CONFIG & CSS
# ═════════════════════════════════════════════════════════════════════════
st.set_page_config(page_title="ChannelPRO™ — Partner Revenue Optimizer", page_icon="📋", layout="wide")
inject_css()

# ═════════════════════════════════════════════════════════════════════════
# LOGIN SCREEN  (delegated to utils.auth)
# ═════════════════════════════════════════════════════════════════════════
handle_login()  # renders login form & calls st.stop() when not authenticated

# ═════════════════════════════════════════════════════════════════════════
# LOGGED IN — INIT
# ═════════════════════════════════════════════════════════════════════════
is_admin = st.session_state.get("auth_role") == "admin"
active_tenant = st.session_state.get("active_tenant")
if active_tenant:
    _tenant_dir(active_tenant)
    _init_criteria()
    if "client_info" not in st.session_state:
        cp = _client_path()
        if cp.exists():
            try: st.session_state["client_info"]=json.loads(cp.read_text())
            except: st.session_state["client_info"]={}
        else: st.session_state["client_info"]={}
if "current_page" not in st.session_state:
    st.session_state["current_page"] = "Client Intake"


# ═════════════════════════════════════════════════════════════════════════
# SIDEBAR (with clickable partner list + PAM filter)
# ═════════════════════════════════════════════════════════════════════════
CLIENT_PAGES = ["Client Intake","Step 1 — Scoring Criteria","Step 2 — Score a Partner","Step 3 — Partner Assessment","Step 4 — Partner Classification","Import Data","Partner List","Ask ChannelPRO™","Break-even — Program Costs","Break-even — Detailed Analysis","Revenue Recovery","User Guide"]
ADMIN_PAGES = CLIENT_PAGES + ["Admin — Manage Users","Admin — All Clients"]

# Pages that require admin role — shown in sidebar for all users ("tease")
# but render a professional placeholder when a client user clicks them.
RESTRICTED_PAGES = {"Step 1 — Scoring Criteria", "Step 4 — Partner Classification"}

with st.sidebar:
    _logo()
    st.markdown("**ChannelPRO™** — Partner Revenue Optimizer")
    st.markdown("---")
    st.markdown(f"👤 **{st.session_state.get('auth_display','User')}**")
    if is_admin:
        st.markdown('<span class="tenant-badge">ADMIN</span>', unsafe_allow_html=True)
    elif active_tenant:
        st.markdown(f'<span class="tenant-badge">{active_tenant}</span>', unsafe_allow_html=True)
    if st.button("🚪 Sign Out", use_container_width=True, type="primary"):
        for k in list(st.session_state.keys()): del st.session_state[k]
        st.rerun()
    st.markdown("---")

    if is_admin:
        tenants = _all_tenants()
        if tenants:
            cur_idx = tenants.index(active_tenant) if active_tenant in tenants else 0
            selected = st.selectbox("🏢 Active Client", tenants, index=cur_idx, key="tenant_sel")
            if selected != active_tenant:
                st.session_state["active_tenant"] = selected
                for k in ["criteria","client_info","_p1_saved","_ci_saved","_p2_submitted","_q_saved","q_config"]:
                    st.session_state.pop(k, None)
                st.rerun()
        else:
            st.info("No clients yet. Create one in Manage Users.")
        st.markdown("---")

    pages = ADMIN_PAGES if is_admin else CLIENT_PAGES
    page = st.radio("Navigate", pages,
        index=pages.index(st.session_state["current_page"]) if st.session_state["current_page"] in pages else 0,
        key="nav_radio", label_visibility="collapsed")
    st.session_state["current_page"] = page
    st.markdown("---")

    chosen_cat = "All Metrics"
    if page not in ("Client Intake","Step 3 — Partner Assessment","Step 4 — Partner Classification","Import Data","Partner List","Ask ChannelPRO™","Break-even — Program Costs","Break-even — Detailed Analysis","Revenue Recovery","Admin — Manage Users","Admin — All Clients"):
        cat_labels=["All Metrics"]+[f"{c['icon']}  {c['label']}" for c in CATEGORIES]
        chosen_cat=st.radio("Category",cat_labels,index=0,label_visibility="collapsed")
    st.markdown("---")

    if active_tenant:
        if _save_path().exists(): st.success("✅ Criteria saved")
        else: st.info("ℹ️ Complete Step 1 first")
        en=_enabled()
        st.metric("Active Metrics",len(en))
        partners=_load_partners()
        # Clickable partner count → expandable list with delete
        mp_limit = _max_partners()
        limit_lbl = f" / {mp_limit}" if mp_limit else ""
        with st.expander(f"📋 Partners Scored: **{len(partners)}{limit_lbl}**"):
            if partners:
                # PAM filter
                pam_names = sorted(set(p.get("pam_name","").strip() for p in partners if p.get("pam_name","").strip()))
                if pam_names:
                    pam_filter = st.selectbox("Filter by PAM", ["All"] + pam_names, key="sb_pam_filter")
                    if pam_filter != "All":
                        partners_show = [p for p in partners if p.get("pam_name","").strip() == pam_filter]
                    else:
                        partners_show = partners
                else:
                    partners_show = partners
                for p in sorted(partners_show, key=lambda x: x.get("partner_name","")):
                    pn = p.get("partner_name","")
                    try: pct = float(p.get("percentage",0) or 0)
                    except: pct = 0
                    gl, gc = _grade(pct)
                    c1, c2, c3 = st.columns([3,1,1])
                    with c1:
                        if st.button(f"📋 {pn}", key=f"sb_view_{pn}", help=f"View scorecard for {pn}", use_container_width=True):
                            st.session_state["_view_partner"] = pn
                            st.session_state["current_page"] = "Step 2 — Score a Partner"
                            st.rerun()
                    with c2: st.markdown(f"<span style='color:{gc};font-weight:700;font-size:.85rem'>{gl}</span>", unsafe_allow_html=True)
                    with c3:
                        if st.button("🗑️", key=f"sb_del_{pn}", help=f"Delete {pn}"):
                            _delete_partner(pn)
                            st.rerun()
                st.markdown("---")
                if st.button("🗑️ Delete All Partners", key="sb_del_all", type="primary", use_container_width=True):
                    st.session_state["_confirm_delete_all"] = True
                if st.session_state.get("_confirm_delete_all"):
                    st.warning("⚠️ This will permanently delete **all** partners. This cannot be undone.")
                    dc1, dc2 = st.columns(2)
                    with dc1:
                        if st.button("✅ Yes, delete all", key="sb_del_all_confirm", use_container_width=True):
                            all_partners = _load_partners()
                            for p in all_partners:
                                _delete_partner(p.get("partner_name",""))
                            st.session_state["_confirm_delete_all"] = False
                            st.rerun()
                    with dc2:
                        if st.button("❌ Cancel", key="sb_del_all_cancel", use_container_width=True):
                            st.session_state["_confirm_delete_all"] = False
                            st.rerun()
            else:
                st.caption("No partners scored yet.")

if chosen_cat=="All Metrics": visible_metrics=SCORECARD_METRICS
else:
    cn=chosen_cat.split("  ",1)[-1]
    ck=next(c["keys"] for c in CATEGORIES if c["label"]==cn)
    visible_metrics=[METRICS_BY_KEY[k] for k in ck]

if not active_tenant and page not in ("Admin — Manage Users","Admin — All Clients"):
    _brand()
    st.warning("No client selected. Use **Admin → Manage Users** to create a client account first.")
    st.stop()

# ── RBAC: "Tease" gate for restricted pages ──────────────────────────
if not is_admin and page in RESTRICTED_PAGES:
    _brand()
    st.markdown(
        '<div style="background:linear-gradient(135deg,#1e2a3a,#2c3e56);border-radius:14px;'
        'padding:40px 36px;margin:30px 0;text-align:center;color:#fff;">'
        '<div style="font-size:2rem;margin-bottom:12px;">🔒</div>'
        '<div style="font-size:1.25rem;font-weight:700;margin-bottom:8px;">Premium Feature</div>'
        '<div style="font-size:.95rem;opacity:.85;line-height:1.6;max-width:520px;margin:0 auto;">'
        'This feature is reserved for full strategic engagements.<br>'
        'Contact your <b>York Group</b> consultant to unlock.</div></div>',
        unsafe_allow_html=True,
    )
    st.stop()


# ═════════════════════════════════════════════════════════════════════════
# CLIENT INTAKE
# ═════════════════════════════════════════════════════════════════════════
if page=="Client Intake":
    _brand()
    st.markdown("""<div class="info-box">
    The <b>Partner Revenue Optimizer</b> is a structured process that will:
    <ol><li>Right-size the margins you provide to your partners, freeing up significant cash flow and revenues; and</li>
    <li>Lay the foundation for targeted partner marketing programs to drive more revenues.</li></ol>
    <p>An experienced channel consultant from <b>The York Group</b> will guide you through the process.
    Each metric is rated <b>1–5</b> (5 = best).</p></div>""",unsafe_allow_html=True)
    if st.session_state.get("_ci_saved"):
        st.markdown('<div class="toast">✅ Client information saved</div>',unsafe_allow_html=True); st.session_state["_ci_saved"]=False
    ci=st.session_state.get("client_info",{})
    with st.form("ci_form"):
        st.markdown('<div class="sec-head">📇 Client Contact Information</div>',unsafe_allow_html=True)
        c1,c2=st.columns(2)
        with c1:
            ci_name=st.text_input("Client name",value=ci.get("client_name",""),key="ci_name")
            ci_url=st.text_input("URL",value=ci.get("url",""),key="ci_url")
            saved_country=ci.get("country","")
            ci_country=st.selectbox("Country",COUNTRIES,index=COUNTRIES.index(saved_country) if saved_country in COUNTRIES else 0,key="ci_country")
            ci_phone=st.text_input("Primary phone",value=ci.get("phone",""),key="ci_phone")
        with c2:
            ci_pm=st.text_input("Client project manager",value=ci.get("project_manager",""),key="ci_pm")
            ci_city=st.text_input("City",value=ci.get("city",""),key="ci_city")
            ci_email=st.text_input("Primary contact email",value=ci.get("email",""),key="ci_email")
            ci_logo_url=st.text_input("Company logo URL",value=ci.get("logo_url",""),key="ci_logo_url",placeholder="https://example.com/logo.png")

        st.markdown('<div class="sec-head">🏢 Client Business Information</div>',unsafe_allow_html=True)
        st.markdown("**Company size — Number of employees**")
        sz_opts=["<100","100-200","200-500","500-1,000","1,000-5,000",">5,000"]; saved_sz=ci.get("company_size",[]); sz_cols=st.columns(len(sz_opts)); sz_sel=[]
        for i,o in enumerate(sz_opts):
            with sz_cols[i]:
                if st.checkbox(o,value=o in saved_sz,key=f"ci_sz_{i}"): sz_sel.append(o)
        st.markdown("**Verticals**")
        v_opts=["Manufacturing","Automotive","Health care","Financial services","Retail","Government","Education","Media and entertainment","Professional services","Life sciences, pharmaceuticals","High-tech, electronics, communications, telecom","None - horizontal solution"]; saved_v=ci.get("verticals",[]); vc=st.columns(3); v_sel=[]
        for i,o in enumerate(v_opts):
            with vc[i%3]:
                if st.checkbox(o,value=o in saved_v,key=f"ci_v_{i}"): v_sel.append(o)
        other_v=st.text_input("Other verticals",value=ci.get("other_verticals",""),key="ci_ov")
        st.markdown("**Solution delivery**")
        d_opts=["On-premise","SaaS/PaaS","IaaS/VM","Device (HW+SW)"]; saved_d=ci.get("solution_delivery",[]); dc=st.columns(len(d_opts)); d_sel=[]
        for i,o in enumerate(d_opts):
            with dc[i]:
                if st.checkbox(o,value=o in saved_d,key=f"ci_d_{i}"): d_sel.append(o)

        st.markdown('<div class="sec-head">🎯 Target Customers</div>',unsafe_allow_html=True)
        st.markdown("**Target customer size — Number of employees**")
        tc_opts=["<100","100-200","200-500","500-1,000","1,000-5,000",">5,000"]; saved_tc=ci.get("target_company_size",[]); tc_cols=st.columns(len(tc_opts)); tc_sel=[]
        for i,o in enumerate(tc_opts):
            with tc_cols[i]:
                if st.checkbox(o,value=o in saved_tc,key=f"ci_tc_{i}"): tc_sel.append(o)
        st.markdown("**Average first-year transaction value**")
        t_opts=["Under $1,000","$1,000–$10,000","$10,000–$50,000","$50,000–$100,000","More than $100,000"]; saved_t=ci.get("avg_transaction_value","")
        txn=st.selectbox("Select range",t_opts,index=t_opts.index(saved_t) if saved_t in t_opts else 0,key="ci_txn")

        st.markdown('<div class="sec-head">📊 Channel Information</div>',unsafe_allow_html=True)
        st.markdown("**Services as % of license/subscription**")
        s_opts=["No services","<20%","20–50%","50–200%",">200%"]; saved_s=ci.get("services_pct","")
        svc=st.selectbox("Select range",s_opts,index=s_opts.index(saved_s) if saved_s in s_opts else 0,key="ci_svc")
        svc_c=st.text_input("Comments",value=ci.get("services_comments",""),key="ci_svc_c")
        st.markdown("**How many resellers/channel partners do you have?**")
        p_opts=["<100","100-200","200-500","500-1,000","1,000-5,000",">5,000"]; saved_p=ci.get("partner_count","")
        pc=st.selectbox("Select range",p_opts,index=p_opts.index(saved_p) if saved_p in p_opts else 0,key="ci_pc")
        st.markdown("**% revenues from indirect channels**")
        i_opts=["<10%","10-30%","30-50%",">50%"]; saved_i=ci.get("indirect_revenue_pct","")
        ind=st.selectbox("Select range",i_opts,index=i_opts.index(saved_i) if saved_i in i_opts else 0,key="ci_ind")
        st.markdown("**Discounts to partners** *(Click all that apply)*")
        disc_opts=["<15%","15-30%","30-50%",">60%","Other"]; saved_disc=ci.get("discounts",[]); disc_c=st.columns(len(disc_opts)); disc_sel=[]
        for i,o in enumerate(disc_opts):
            with disc_c[i]:
                if st.checkbox(o,value=o in saved_disc,key=f"ci_disc_{i}"): disc_sel.append(o)
        st.markdown("**Partner designations**")
        desig=st.text_input("Comma-separated, e.g. gold, silver, bronze",value=ci.get("partner_designations",""),key="ci_desig")
        st.markdown("---")
        _,cr=st.columns([3,1])
        with cr: ci_sub=st.form_submit_button("Next →  Step 1",use_container_width=True,type="primary")
    if ci_sub:
        st.session_state["client_info"]={"client_name":ci_name,"project_manager":ci_pm,"url":ci_url,"city":ci_city,"country":ci_country,"email":ci_email,"phone":ci_phone,"logo_url":ci_logo_url,"company_size":sz_sel,"verticals":v_sel,"other_verticals":other_v,"solution_delivery":d_sel,"target_company_size":tc_sel,"avg_transaction_value":txn,"services_pct":svc,"services_comments":svc_c,"partner_count":pc,"indirect_revenue_pct":ind,"discounts":disc_sel,"partner_designations":desig}
        _client_path().write_text(json.dumps(st.session_state["client_info"],indent=2))
        st.session_state["_ci_saved"]=True; st.session_state["current_page"]="Step 1 — Scoring Criteria"; st.rerun()


# ═════════════════════════════════════════════════════════════════════════
# STEP 1 — SCORING CRITERIA
# ═════════════════════════════════════════════════════════════════════════
elif page=="Step 1 — Scoring Criteria":
    _brand()
    st.markdown("## Step 1 — Define Scoring Criteria")
    st.markdown("Configure **1–5** thresholds. Toggle metrics on/off. Changes apply retroactively to all scored partners.")
    if st.session_state.get("_p1_saved"):
        st.markdown('<div class="toast">✅ Criteria saved — all partners re-scored</div>',unsafe_allow_html=True); st.session_state["_p1_saved"]=False
    with st.form("p1_form"):
        for m in visible_metrics:
            mk=m["key"]; cr=st.session_state["criteria"].get(mk,{}); iq=m["type"]=="quantitative"
            if not cr: continue  # Skip metrics not yet in criteria
            ie=cr.get("enabled",True)
            tt='<span class="tag tag-q">Quantitative</span>' if iq else '<span class="tag tag-ql">Qualitative</span>'
            dt=f'<span class="tag {"tag-hi" if m["direction"]=="higher_is_better" else "tag-lo"}">{"↑ Higher" if m["direction"]=="higher_is_better" else "↓ Lower"} is better</span>'
            xt='' if ie else '<span class="tag tag-del">EXCLUDED</span>'
            dc="" if ie else " mc-off"
            ud=f' ({m["unit"]})' if m.get("unit") else ""
            st.markdown(f'<div class="mc{dc}"><span class="mname">{m["id"]}. {m["name"]}</span>{tt}{dt}{xt}<div class="mexpl">{m["explanation"]}</div></div>',unsafe_allow_html=True)
            st.checkbox("Include in scoring",value=ie,key=f"p1_{mk}_en")
            if iq:
                cols=st.columns(5)
                for idx,s in enumerate(("1","2","3","4","5")):
                    with cols[idx]:
                        st.markdown(f'<div class="sb sb{s}">{s}</div>',unsafe_allow_html=True)
                        st.text_input(f"Min{ud}",value=cr["ranges"][s]["min"],key=f"p1_{mk}_s{s}_min",placeholder="No min" if s=="1" else "")
                        st.text_input(f"Max{ud}",value=cr["ranges"][s]["max"],key=f"p1_{mk}_s{s}_max",placeholder="No cap" if s=="5" else "")
            else:
                cols=st.columns(5)
                for idx,s in enumerate(("1","2","3","4","5")):
                    with cols[idx]:
                        st.markdown(f'<div class="sb sb{s}">{s}</div>',unsafe_allow_html=True)
                        st.text_area("desc",value=cr["descriptors"][s],key=f"p1_{mk}_s{s}_desc",height=100,label_visibility="collapsed")
        st.markdown("---")
        st.markdown('<style>[data-testid="stForm"] [data-testid="stFormSubmitButton"]:last-child button{background:#dc4040!important;border-color:#dc4040!important;color:#fff!important;font-weight:800!important}</style>',unsafe_allow_html=True)
        _,cm,cr=st.columns([2,1,1])
        with cm: p1s=st.form_submit_button("💾  Save Criteria",use_container_width=True)
        with cr: p1n=st.form_submit_button("Next →  Step 2",use_container_width=True,type="primary")
    if p1s or p1n:
        _save_criteria(); st.session_state["_p1_saved"]=True
        if p1n: st.session_state["current_page"]="Step 2 — Score a Partner"
        st.rerun()

    # ── Auto-calculate from data ──
    if _raw_path().exists():
        st.markdown("---")
        st.markdown("#### 📐 Auto-Calculate from Partner Data")
        st.caption("Analyze your imported partner data and set quintile-based ranges automatically. Qualitative metrics are not affected.")

        if st.session_state.get("_bench_result_p1"):
            res = st.session_state.pop("_bench_result_p1")
            if res["updated"]:
                st.success(f"Benchmarks recalculated — **{len(res['updated'])}** metric(s) updated. All partners re-scored.")
            else:
                st.info("No quantitative metrics were changed.")

        _, _, bc = st.columns([2, 2, 1])
        with bc:
            if st.button("📐  Recalculate Benchmarks", use_container_width=True, key="p1_bench"):
                with st.spinner("Analyzing partner data distributions..."):
                    result = _recalculate_benchmarks()
                st.session_state["_bench_result_p1"] = result
                st.rerun()


# ═════════════════════════════════════════════════════════════════════════
# STEP 2 — SCORE A PARTNER
# ═════════════════════════════════════════════════════════════════════════
elif page=="Step 2 — Score a Partner":
    _brand()
    st.markdown("## Step 2 — Score a Partner")
    if not _save_path().exists():
        st.warning("⚠️ Complete **Step 1** first."); st.stop()
    st.session_state["criteria"]=json.loads(_save_path().read_text())
    _ensure_criteria_complete()
    cr=st.session_state["criteria"]; em=_enabled(); mx=len(em)*5
    # Form version counter — incremented on submit to clear all fields
    if "p2_ver" not in st.session_state: st.session_state["p2_ver"]=0
    fv=st.session_state["p2_ver"]
    # Edit mode: if a partner was clicked in sidebar or assessment
    # Transfer _view_partner to persistent _editing_partner
    if "_view_partner" in st.session_state:
        st.session_state["_editing_partner"] = st.session_state.pop("_view_partner")
    view_raw = None
    is_edit = False
    editing_pn = st.session_state.get("_editing_partner")
    if editing_pn:
        raw_all = _load_raw()
        view_raw = next((r for r in raw_all if r.get("partner_name") == editing_pn), None)
        if view_raw:
            is_edit = True
            col_edit_info, col_edit_cancel = st.columns([5, 1])
            with col_edit_info:
                st.info(f"✏️ Editing scorecard for **{editing_pn}** — make changes and click **Save Changes**")
            with col_edit_cancel:
                if st.button("✖ Cancel Edit", key="p2_cancel_edit"):
                    st.session_state.pop("_editing_partner", None)
                    st.rerun()
        else:
            st.session_state.pop("_editing_partner", None)
    if st.session_state.get("_p2_submitted"):
        st.markdown('<div class="toast">✅ Partner submitted & saved. Ready for next partner.</div>',unsafe_allow_html=True); st.session_state["_p2_submitted"]=False
    if st.session_state.get("_p2_updated"):
        st.markdown('<div class="toast">✅ Partner scorecard updated successfully.</div>',unsafe_allow_html=True); st.session_state["_p2_updated"]=False
    st.markdown(f"Total out of **{mx}** ({len(em)} metrics × 5).")
    st.markdown('<div class="partner-hdr">Partner details</div>',unsafe_allow_html=True)
    tiers=_tiers(); t_opts=["Please choose..."]+tiers if tiers else ["Please choose...","(Set tiers in Client Intake)"]
    pc1,pc2,pc3,pc4=st.columns(4)
    with pc1: pn=st.text_input("Partner company name",key=f"p2_pn_{fv}",placeholder="e.g. ABC reseller",value=view_raw.get("partner_name","") if view_raw else "")
    with pc2: p_yr=st.text_input("Year became partner",key=f"p2_py_{fv}",placeholder="e.g. 2020",value=view_raw.get("partner_year","") if view_raw else "")
    with pc3:
        tier_val = view_raw.get("partner_tier","") if view_raw else ""
        pt=st.selectbox("Tier",t_opts,index=t_opts.index(tier_val) if tier_val in t_opts else 0,key=f"p2_pt_{fv}")
    with pc4:
        p_disc=st.text_input("Partner Discount",key=f"p2_disc_{fv}",placeholder="e.g. 20%",value=view_raw.get("partner_discount","") if view_raw else "")
    pc5,pc6=st.columns(2)
    with pc5: pcity=st.text_input("City",key=f"p2_city_{fv}",placeholder="e.g. Paris",value=view_raw.get("partner_city","") if view_raw else "")
    with pc6:
        country_val = view_raw.get("partner_country","") if view_raw else ""
        pcountry=st.selectbox("Country",COUNTRIES,index=COUNTRIES.index(country_val) if country_val in COUNTRIES else 0,key=f"p2_country_{fv}")
    pam1,pam2=st.columns(2)
    with pam1: pam_name=st.text_input("Partner Account Manager (PAM) name",key=f"p2_pam_name_{fv}",placeholder="e.g. Jane Smith",value=view_raw.get("pam_name","") if view_raw else "")
    with pam2: pam_email=st.text_input("PAM email",key=f"p2_pam_email_{fv}",placeholder="e.g. jane@company.com",value=view_raw.get("pam_email","") if view_raw else "")
    st.markdown("---")
    if chosen_cat=="All Metrics": ve=em
    else:
        cn=chosen_cat.split("  ",1)[-1]; ck=next(c["keys"] for c in CATEGORIES if c["label"]==cn); ve=[m for m in em if m["key"] in ck]
    for m in ve:
        mk=m["key"]; mc=cr.get(mk,{}); iq=m["type"]=="quantitative"
        if not mc: continue
        tt='<span class="tag tag-q">Quantitative</span>' if iq else '<span class="tag tag-ql">Qualitative</span>'
        dt=f'<span class="tag {"tag-hi" if m["direction"]=="higher_is_better" else "tag-lo"}">{"↑ Higher" if m["direction"]=="higher_is_better" else "↓ Lower"} is better</span>'
        st.markdown(f'<div class="mc"><span class="mname">{m["id"]}. {m["name"]}</span>{tt}{dt}<div class="mexpl">{m["explanation"]}</div></div>',unsafe_allow_html=True)
        view_val = view_raw.get(f"raw_{mk}","") if view_raw else ""
        if iq:
            u=mc.get("unit","") or ""
            hints=[]
            for s in("1","2","3","4","5"):
                r=mc["ranges"][s]; lo,hi=r["min"],r["max"]
                if lo and hi: hints.append(f"<b>{s}</b>: {lo}–{hi}")
                elif lo and not hi: hints.append(f"<b>{s}</b>: ≥{lo}")
                elif not lo and hi: hints.append(f"<b>{s}</b>: ≤{hi}")
            if hints and is_admin: st.markdown(f'<div class="hint-row">Ranges ({u}): {" &nbsp;·&nbsp; ".join(hints)}</div>',unsafe_allow_html=True)
            ic,sc_c=st.columns([4,1])
            with ic: pv=st.text_input(f"Value ({u})",key=f"p2_{mk}_{fv}",placeholder=f"Enter number ({u})",label_visibility="collapsed",value=str(view_val) if view_val else "")
            scr=calc_score(mk,pv)
        else:
            opts=["— Select —"]+[f"({s}) {mc['descriptors'][s]}" for s in("1","2","3","4","5")]
            # Pre-select for edit mode — match descriptor exactly
            view_idx = 0
            if view_val:
                for oi, o in enumerate(opts):
                    if oi == 0: continue  # skip "— Select —"
                    desc = re.sub(r"^\(\d\)\s*", "", o)
                    if desc == view_val:
                        view_idx = oi; break
                else:
                    # Fallback: substring match if exact fails
                    for oi, o in enumerate(opts):
                        if oi > 0 and view_val in o: view_idx = oi; break
            ic,sc_c=st.columns([4,1])
            with ic: pv=st.selectbox("Level",opts,index=view_idx,key=f"p2_{mk}_{fv}",label_visibility="collapsed")
            if pv and pv!="— Select —": raw_d=re.sub(r"^\(\d\)\s*","",pv); scr=calc_score(mk,raw_d)
            else: scr=None
        with sc_c:
            if scr: st.markdown(f'<div class="live-score" style="background:{SC[scr]}">{scr}</div>',unsafe_allow_html=True)
            else: st.markdown('<div class="live-score" style="background:#ccc">—</div>',unsafe_allow_html=True)
    # Summary
    st.markdown("---")
    full={}; raw_vals={}
    for m in em:
        mk=m["key"]; pv=st.session_state.get(f"p2_{mk}_{fv}","")
        if not pv or pv=="— Select —":
            full[mk]=None; raw_vals[mk]=None
        elif m["type"]=="qualitative" and isinstance(pv,str) and pv.startswith("("):
            raw_d=re.sub(r"^\(\d\)\s*","",pv); full[mk]=calc_score(mk,raw_d); raw_vals[mk]=raw_d
        else:
            full[mk]=calc_score(mk,pv); raw_vals[mk]=pv
    si={k:v for k,v in full.items() if v is not None}; total=sum(si.values()); sn=len(si); mp=sn*5
    pct=(total/mp*100) if mp else 0; gl,gc=_grade(pct); pname=st.session_state.get(f"p2_pn_{fv}","Partner") or "Partner"
    st.markdown(f"### Live Summary — {pname}")
    c1,c2,c3,c4=st.columns(4)
    with c1: st.markdown(f'<div class="sum-card"><div class="sum-big">{total}</div><div class="sum-lbl">Total</div></div>',unsafe_allow_html=True)
    with c2: st.markdown(f'<div class="sum-card"><div class="sum-big">{sn}/{len(em)}</div><div class="sum-lbl">Scored</div></div>',unsafe_allow_html=True)
    with c3: st.markdown(f'<div class="sum-card"><div class="sum-big">{pct:.1f}%</div><div class="sum-lbl">Percentage</div></div>',unsafe_allow_html=True)
    with c4: st.markdown(f'<div class="sum-card"><div class="sum-big" style="color:{gc}">{gl}</div><div class="sum-lbl">Grade</div></div>',unsafe_allow_html=True)
    st.markdown("---")
    # Build row + raw dicts (shared by both submit and save)
    def _build_row_raw():
        row={"partner_name":pname,"partner_year":st.session_state.get(f"p2_py_{fv}",""),
             "partner_tier":st.session_state.get(f"p2_pt_{fv}",""),
             "partner_discount":st.session_state.get(f"p2_disc_{fv}",""),
             "partner_city":st.session_state.get(f"p2_city_{fv}",""),
             "partner_country":st.session_state.get(f"p2_country_{fv}",""),
             "pam_name":st.session_state.get(f"p2_pam_name_{fv}",""),
             "pam_email":st.session_state.get(f"p2_pam_email_{fv}",""),
             "total_score":total,"max_possible":mp,"percentage":round(pct,1)}
        for m in em: row[m["key"]]=full.get(m["key"],"")
        raw_dict={"partner_name":pname,"partner_year":st.session_state.get(f"p2_py_{fv}",""),
                  "partner_tier":st.session_state.get(f"p2_pt_{fv}",""),
                  "partner_discount":st.session_state.get(f"p2_disc_{fv}",""),
                  "partner_city":st.session_state.get(f"p2_city_{fv}",""),
                  "partner_country":st.session_state.get(f"p2_country_{fv}",""),
                  "pam_name":st.session_state.get(f"p2_pam_name_{fv}",""),
                  "pam_email":st.session_state.get(f"p2_pam_email_{fv}","")}
        for m in em: raw_dict[f"raw_{m['key']}"]=raw_vals.get(m["key"])
        return row, raw_dict
    if is_edit:
        # Edit mode — update existing partner
        _,_,save_col=st.columns([2,2,1])
        with save_col:
            if st.button("💾  Save Changes",use_container_width=True,type="primary"):
                if not pname or pname=="Partner":
                    st.error("Partner name is missing.")
                else:
                    row, raw_dict = _build_row_raw()
                    _upsert_partner(row, raw_dict)
                    st.session_state.pop("_editing_partner", None)
                    st.session_state["_p2_updated"]=True; st.rerun()
    else:
        # New partner mode
        _,_,submit_col=st.columns([2,2,1])
        with submit_col:
            if st.button("✅  Submit & Start New Partner",use_container_width=True,type="primary"):
                if not pname or pname=="Partner":
                    st.error("Enter a partner name first.")
                elif _partner_exists(pname):
                    st.error(f"A partner named **{pname}** already exists. Use a unique name.")
                elif _max_partners() and _partner_count() >= _max_partners():
                    st.error(f"Partner limit reached (**{_max_partners()}**). Contact your admin to increase the limit.")
                else:
                    row, raw_dict = _build_row_raw()
                    _append_partner(row, raw_dict)
                    st.session_state["p2_ver"]=fv+1
                    st.session_state["_p2_submitted"]=True; st.rerun()


# ═════════════════════════════════════════════════════════════════════════
# STEP 3 — PARTNER ASSESSMENT
# ═════════════════════════════════════════════════════════════════════════
elif page=="Step 3 — Partner Assessment":
    _brand(); st.markdown("## Step 3 — Partner Assessment")
    partners=_load_partners(); em=_enabled()
    if not partners: st.info("No partners scored yet. Complete **Step 2**."); st.stop()

    # ── Build DataFrame for display ──
    rows_data = []
    for p in partners:
        row = {
            "Partner": p.get("partner_name",""),
            "Country": p.get("partner_country",""),
            "Tier": p.get("partner_tier",""),
            "Discount": p.get("partner_discount",""),
            "PAM": p.get("pam_name",""),
        }
        for m in em:
            try: v = int(p.get(m["key"],"") or 0)
            except: v = 0
            row[m["name"]] = v if v else 0
        try: row["Total"] = int(p.get("total_score",0) or 0)
        except: row["Total"] = 0
        try: row["Pct"] = round(float(p.get("percentage",0) or 0), 1)
        except: row["Pct"] = 0.0
        gl, gc = _grade(row["Pct"])
        row["Grade"] = gl
        rows_data.append(row)

    df_grid = pd.DataFrame(rows_data)
    # Ensure metric columns are numeric
    for m in em:
        df_grid[m["name"]] = pd.to_numeric(df_grid[m["name"]], errors="coerce").fillna(0).astype(int)
    df_grid["Total"] = pd.to_numeric(df_grid["Total"], errors="coerce").fillna(0).astype(int)
    df_grid["Pct"] = pd.to_numeric(df_grid["Pct"], errors="coerce").fillna(0.0)

    st.markdown(f"**{len(df_grid)}** partners · **{len(em)}** metrics")

    if HAS_AGGRID:
        # ══════════════════════════════════════════════════════════════
        # AG-GRID — PROFESSIONAL ALPINE-THEMED TABLE
        # ══════════════════════════════════════════════════════════════
        st.markdown(
            '<p style="font-size:0.88rem;color:#475569;margin-bottom:12px;">'
            'Use the filter icons in each column header to filter. '
            'Click any header to sort. Metrics are grouped by category \u2014 '
            'scroll horizontally to see all groups. '
            '<b>Partner</b> is pinned left; '
            '<b>Total / Score% / Grade</b> are pinned right.</p>',
            unsafe_allow_html=True,
        )

        grid_response = display_styled_assessment_table(df_grid, em)

        # --- Handle row selection → open scorecard ---
        selected_rows = grid_response.get("selected_rows", None)
        if selected_rows is not None and len(selected_rows) > 0:
            if isinstance(selected_rows, pd.DataFrame):
                sel_pn = selected_rows.iloc[0].get("Partner", "")
            else:
                sel_pn = selected_rows[0].get("Partner", "")
            if sel_pn:
                st.markdown("---")
                oc1, oc2 = st.columns([3, 1])
                with oc1:
                    st.markdown(f"**Selected:** {sel_pn}")
                with oc2:
                    if st.button("✏️  Open Scorecard", use_container_width=True, type="primary", key="p3_open_btn"):
                        st.session_state["_view_partner"] = sel_pn
                        st.session_state["current_page"] = "Step 2 — Score a Partner"
                        st.rerun()

        # --- Clickable partner access (always available as fallback) ---
        st.markdown("---")
        st.markdown("#### 📋 Open Partner Scorecard")
        partner_names_sorted = sorted(df_grid["Partner"].tolist())
        if partner_names_sorted:
            oc1, oc2 = st.columns([3, 1])
            with oc1:
                open_pn = st.selectbox("Select a partner to view/edit their scorecard", partner_names_sorted, key="p3_open_partner_sel")
            with oc2:
                st.markdown("<br>", unsafe_allow_html=True)
                if st.button("✏️  Open Scorecard", use_container_width=True, type="primary", key="p3_open_btn2"):
                    st.session_state["_view_partner"] = open_pn
                    st.session_state["current_page"] = "Step 2 — Score a Partner"
                    st.rerun()

        # --- Downloads (use the full DataFrame — filtered by AgGrid state is the displayed data) ---
        st.markdown("---")
        dl1, dl2 = st.columns(2)
        with dl1:
            xb = _gen_xlsx(partners, em)
            if xb:
                st.download_button("⬇️  Download Excel", xb, "Partner_Assessment.xlsx",
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", type="primary")
        with dl2:
            csv_buf = df_grid.to_csv(index=False)
            st.download_button("⬇️  Download CSV", csv_buf, "Partner_Assessment.csv", "text/csv")

    else:
        # ══════════════════════════════════════════════════════════════
        # FALLBACK: original HTML heatmap table (when st_aggrid not installed)
        # ══════════════════════════════════════════════════════════════
        st.warning("📦 Install `streamlit-aggrid` for interactive column filtering: `pip install streamlit-aggrid`")

        # ── Quick filters row ──
        f1, f2, f3, f4 = st.columns([2, 2, 2, 2])
        with f1:
            search_q = st.text_input("🔍 Search partner", key="p3_search", placeholder="Type to search...")
        with f2:
            pam_names = sorted(set(p.get("pam_name","").strip() for p in partners if p.get("pam_name","").strip()))
            pam_f = st.selectbox("Filter by PAM", ["All PAMs"] + pam_names, key="p3_pam_filter")
        with f3:
            sort_opts = ["Score (highest first)", "Score (lowest first)", "Partner (A–Z)", "Partner (Z–A)", "PAM (A–Z)"]
            sort_by = st.selectbox("Sort by", sort_opts, key="p3_sort")
        with f4:
            metric_filter_opts = ["All metrics"] + [m["name"] for m in em]
            metric_filter = st.selectbox("Filter by metric score", metric_filter_opts, key="p3_metric_filter")

        # ── Apply filters ──
        ps = list(partners)
        if search_q:
            sq = search_q.strip().lower()
            ps = [p for p in ps if sq in p.get("partner_name","").lower() or sq in p.get("pam_name","").lower() or sq in p.get("partner_country","").lower()]
        if pam_f != "All PAMs":
            ps = [p for p in ps if p.get("pam_name","").strip() == pam_f]
        filter_mk = None
        if metric_filter != "All metrics":
            filter_mk = next((m["key"] for m in em if m["name"] == metric_filter), None)

        # ── Sort ──
        if sort_by == "Score (highest first)":
            ps = sorted(ps, key=lambda p: -int(p.get("total_score",0) or 0))
        elif sort_by == "Score (lowest first)":
            ps = sorted(ps, key=lambda p: int(p.get("total_score",0) or 0))
        elif sort_by == "Partner (A–Z)":
            ps = sorted(ps, key=lambda p: p.get("partner_name","").lower())
        elif sort_by == "Partner (Z–A)":
            ps = sorted(ps, key=lambda p: p.get("partner_name","").lower(), reverse=True)
        elif sort_by == "PAM (A–Z)":
            ps = sorted(ps, key=lambda p: p.get("pam_name","").lower())
        if filter_mk:
            ps = sorted(ps, key=lambda p: -int(p.get(filter_mk,"") or 0))

        st.markdown(f"**{len(ps)}** partners, **{len(em)}** metrics.")

        # ── Build heatmap table ──
        hdr="<tr><th>Rank</th><th style='text-align:left'>Partner</th><th>Country</th><th>Tier</th><th>PAM</th>"
        for m in em: hdr+=f'<th class="hm-diag" title="{m["name"]}"><div>{m["name"][:25]}</div></th>'
        hdr+="<th>Total</th><th>%</th></tr>"
        rows=""
        for ri,p in enumerate(ps,1):
            pn_display = p.get('partner_name','')
            rows+=f"<tr><td><b>{ri}</b></td><td style='text-align:left;padding-left:10px;white-space:nowrap'>{pn_display}</td><td style='white-space:nowrap'>{p.get('partner_country','')}</td><td>{p.get('partner_tier','')}</td><td style='white-space:nowrap'>{p.get('pam_name','')}</td>"
            for m in em:
                try: v=int(p.get(m["key"],"") or 0)
                except: v=None
                if v and 1<=v<=5: rows+=f'<td class="hm{v}">{v}</td>'
                else: rows+='<td style="color:#ccc">—</td>'
            try: tv=int(p.get("total_score",0) or 0)
            except: tv=0
            try: pv=float(p.get("percentage",0) or 0)
            except: pv=0
            rows+=f'<td class="hm-total">{tv}</td><td class="hm-total">{pv:.1f}%</td></tr>'
        st.markdown(f'<div class="scroll-tbl"><table class="hm-tbl"><thead>{hdr}</thead><tbody>{rows}</tbody></table></div>',unsafe_allow_html=True)

        # ── Clickable partner access ──
        st.markdown("---")
        st.markdown("#### 📋 Open Partner Scorecard")
        partner_names_sorted = [p.get("partner_name","") for p in ps]
        if partner_names_sorted:
            oc1, oc2 = st.columns([3, 1])
            with oc1:
                open_pn = st.selectbox("Select a partner to view/edit their scorecard", partner_names_sorted, key="p3_open_partner")
            with oc2:
                st.markdown("<br>", unsafe_allow_html=True)
                if st.button("✏️  Open Scorecard", use_container_width=True, type="primary", key="p3_open_btn"):
                    st.session_state["_view_partner"] = open_pn
                    st.session_state["current_page"] = "Step 2 — Score a Partner"
                    st.rerun()

        st.markdown("---")
        xb=_gen_xlsx(ps,em)
        if xb: st.download_button("⬇️  Download Excel",xb,"Partner_Assessment.xlsx","application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",type="primary")
        if _csv_path().exists(): st.download_button("⬇️  Download CSV",_csv_path().read_text(),"all_partners.csv","text/csv")


# ═════════════════════════════════════════════════════════════════════════
# STEP 4 — PARTNER CLASSIFICATION (3 quadrants + Long Tail)
# ═════════════════════════════════════════════════════════════════════════
elif page=="Step 4 — Partner Classification":
    _brand(); st.markdown("## Step 4 — Partner Classification")
    partners=_load_partners(); em=_enabled(); em_keys={m["key"] for m in em}
    if not partners: st.info("No partners scored yet. Complete **Step 2** first."); st.stop()
    st.markdown("""<div class="info-box">Classify partners into <b>three quadrants</b> plus <b>Long Tail</b>. Partners that don't match Q1–Q3 criteria are automatically classified as Long Tail.<br>Levels: <b>High</b> ≥ 4, <b>Mid</b> = 3, <b>Low</b> ≤ 2, <b>Any</b> = score &gt; 0. First match wins (1 → 3).</div>""",unsafe_allow_html=True)
    q_config=_load_q_config()
    if "q_config" not in st.session_state: st.session_state["q_config"]=q_config
    if st.session_state.get("_q_saved"):
        st.markdown('<div class="toast">✅ Classification saved & applied</div>',unsafe_allow_html=True); st.session_state["_q_saved"]=False
    st.markdown("### Define Quadrant Criteria")
    metric_options=["(none)"]+[m["name"] for m in SCORECARD_METRICS]; level_options=["high","mid","low","any"]; MAX_C=6
    with st.form("class_form"):
        new_config={}
        for qn in(1,2,3):
            ql,qc=Q_LABELS.get(qn,(f"Q{qn}","#666"))
            st.markdown(f'<div style="display:flex;align-items:center;margin:18px 0 8px;"><div class="q-badge" style="background:{qc}">{qn}</div><b style="font-size:1.05rem">{ql}</b></div>',unsafe_allow_html=True)
            st.caption(Q_DESCS.get(qn,""))
            current=st.session_state["q_config"].get(qn,[])
            crit_list=[]
            for ci in range(MAX_C):
                if ci<len(current): cur_key,cur_level=current[ci]; cur_mn=next((m["name"] for m in SCORECARD_METRICS if m["key"]==cur_key),"(none)"); cur_li=level_options.index(cur_level) if cur_level in level_options else 0
                else: cur_mn="(none)"; cur_li=0
                mc,lc=st.columns([3,1])
                with mc: sel_m=st.selectbox(f"Q{qn}M{ci+1}",metric_options,index=metric_options.index(cur_mn) if cur_mn in metric_options else 0,key=f"q{qn}_m{ci}",label_visibility="collapsed")
                with lc: sel_l=st.selectbox(f"Q{qn}L{ci+1}",level_options,index=cur_li,key=f"q{qn}_l{ci}",label_visibility="collapsed")
                if sel_m!="(none)":
                    mk=next((m["key"] for m in SCORECARD_METRICS if m["name"]==sel_m),None)
                    if mk: crit_list.append((mk,sel_l))
            new_config[qn]=crit_list
        st.markdown(f'<div style="display:flex;align-items:center;margin:18px 0 8px;"><div class="q-badge" style="background:#dc4040">4</div><b style="font-size:1.05rem">Long Tail</b></div>',unsafe_allow_html=True)
        st.caption("Automatically assigned to partners that don't match Q1–Q3 criteria.")
        st.markdown("---"); _,cc=st.columns([3,1])
        with cc: q_sub=st.form_submit_button("💾  Save & Classify",use_container_width=True,type="primary")
    if q_sub:
        st.session_state["q_config"]=new_config; _save_q_config(new_config); st.session_state["_q_saved"]=True; st.rerun()
    st.markdown("---"); st.markdown("### Classification Results")
    ac=st.session_state["q_config"]; classification=classify_partners(partners,ac,em_keys)
    if not classification: st.info("No partners to classify."); st.stop()
    by_q={1:[],2:[],3:[],4:[]}
    for pn,qn in classification.items(): by_q.setdefault(qn,[]).append(pn)
    for qn in(1,2,3,4):
        ql,qc=Q_LABELS.get(qn,(f"Q{qn}","#666")); members=by_q.get(qn,[]); cnt=len(members)
        if qn<=3:
            cp=[]
            for(mk,lvl) in ac.get(qn,[]):
                mn=next((m["name"] for m in SCORECARD_METRICS if m["key"]==mk),mk); cp.append(f"{mn} = <b>{lvl}</b>")
            ch=" &nbsp;·&nbsp; ".join(cp) if cp else "<i>No criteria</i>"
        else:
            ch="<i>Does not match Q1–Q3 criteria</i>"
        ph="".join(f'<span class="q-partner">{n}</span>' for n in members) if members else '<span style="color:#999">None</span>'
        st.markdown(f'<div class="q-card" style="border-color:{qc}20;background:{qc}08;"><div style="display:flex;align-items:center;"><div class="q-badge" style="background:{qc}">{qn}</div><h4 style="margin:0;color:{qc}">{ql}</h4><span style="margin-left:auto;font-size:1.4rem;font-weight:800;color:{qc};font-family:\'JetBrains Mono\',monospace">{cnt}</span></div><div class="q-criteria">Criteria: {ch}</div><div style="margin-top:10px">{ph}</div></div>',unsafe_allow_html=True)
    st.markdown("### Full Classification Table")
    tbl="<table class='hm-tbl'><thead><tr><th>Partner</th><th>Total</th><th>%</th><th>Q</th><th>Classification</th></tr></thead><tbody>"
    sc=sorted(classification.items(),key=lambda x:(x[1],-next((int(p.get("total_score",0) or 0) for p in partners if p.get("partner_name")==x[0]),0)))
    for pn,qn in sc:
        p=next((p for p in partners if p.get("partner_name")==pn),{})
        try: tv=int(p.get("total_score",0) or 0)
        except: tv=0
        try: pv=float(p.get("percentage",0) or 0)
        except: pv=0
        ql,qc=Q_LABELS.get(qn,(f"Q{qn}","#666"))
        tbl+=f'<tr><td style="text-align:left;padding-left:10px;font-weight:600">{pn}</td><td>{tv}</td><td>{pv:.1f}%</td><td><span class="score-pill" style="background:{qc}">{qn}</span></td><td style="color:{qc};font-weight:700">{ql}</td></tr>'
    tbl+="</tbody></table>"
    st.markdown(tbl,unsafe_allow_html=True)
    st.markdown("---")
    # Downloads: Excel, CSV, JSON
    dl1, dl2, dl3 = st.columns(3)
    # Build classification data for export
    class_rows = []
    for pn, qn in sorted(classification.items(), key=lambda x: (x[1], x[0])):
        p = next((p for p in partners if p.get("partner_name") == pn), {})
        try: tv = int(p.get("total_score", 0) or 0)
        except: tv = 0
        try: pv = float(p.get("percentage", 0) or 0)
        except: pv = 0
        ql, _ = Q_LABELS.get(qn, (f"Q{qn}", "#666"))
        class_rows.append({"Partner": pn, "Total Score": tv, "Percentage": round(pv, 1),
                           "Quadrant": qn, "Classification": ql,
                           "Tier": p.get("partner_tier", ""), "PAM": p.get("pam_name", "")})
    with dl1:
        # Excel
        try:
            import openpyxl
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Classification"
            q_fills = {1: PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid"),
                       2: PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
                       3: PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
                       4: PatternFill(start_color="FA7A7A", end_color="FA7A7A", fill_type="solid")}
            hf = PatternFill(start_color="1E2A3A", end_color="1E2A3A", fill_type="solid")
            hfont = Font(color="FFFFFF", bold=True, size=10)
            bdr = Border(left=Side(style="thin", color="CCCCCC"), right=Side(style="thin", color="CCCCCC"),
                         top=Side(style="thin", color="CCCCCC"), bottom=Side(style="thin", color="CCCCCC"))
            headers = ["Partner", "Tier", "PAM", "Total Score", "Percentage", "Quadrant", "Classification"]
            for ci, h in enumerate(headers, 1):
                cell = ws.cell(1, ci, h); cell.fill = hf; cell.font = hfont; cell.border = bdr
                cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions["A"].width = 28; ws.column_dimensions["B"].width = 14
            ws.column_dimensions["C"].width = 22; ws.column_dimensions["G"].width = 24
            for ri, r in enumerate(class_rows, 2):
                ws.cell(ri, 1, r["Partner"]).border = bdr
                ws.cell(ri, 2, r["Tier"]).border = bdr
                ws.cell(ri, 3, r["PAM"]).border = bdr
                ws.cell(ri, 4, r["Total Score"]).border = bdr; ws.cell(ri, 4).alignment = Alignment(horizontal="center")
                pc = ws.cell(ri, 5); pc.value = r["Percentage"] / 100; pc.number_format = "0.0%"; pc.border = bdr; pc.alignment = Alignment(horizontal="center")
                qc = ws.cell(ri, 6, r["Quadrant"]); qc.border = bdr; qc.alignment = Alignment(horizontal="center")
                qc.fill = q_fills.get(r["Quadrant"], PatternFill())
                ws.cell(ri, 7, r["Classification"]).border = bdr; ws.cell(ri, 7).font = Font(bold=True)
            buf = io.BytesIO(); wb.save(buf)
            st.download_button("⬇️  Download Excel", buf.getvalue(), "Partner_Classification.xlsx",
                              "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", type="primary")
        except ImportError:
            st.warning("openpyxl required for Excel export")
    with dl2:
        # CSV
        csv_buf = io.StringIO()
        if class_rows:
            cw = csv.DictWriter(csv_buf, fieldnames=list(class_rows[0].keys()))
            cw.writeheader()
            for r in class_rows: cw.writerow(r)
        st.download_button("⬇️  Download CSV", csv_buf.getvalue(), "Partner_Classification.csv", "text/csv")
    with dl3:
        st.download_button("⬇️  Download JSON", json.dumps(classification, indent=2), "partner_classification.json", "application/json")


# ═════════════════════════════════════════════════════════════════════════
# IMPORT DATA — CSV upload + column mapping → batch partner creation
# ═════════════════════════════════════════════════════════════════════════
elif page=="Import Data":
    _brand(); st.markdown("## Import Partner Data from CSV")
    if not _save_path().exists():
        st.warning("⚠️ Complete **Step 1 — Scoring Criteria** first so metrics are available for mapping."); st.stop()
    st.session_state["criteria"] = json.loads(_save_path().read_text())
    _ensure_criteria_complete()
    cr = st.session_state["criteria"]; em = _enabled()

    st.markdown("""<div class="info-box">
    Upload a CSV exported from your CRM, ERP, or PRM system. Map its columns to ChannelPRO™ scoring metrics
    and import partners in bulk. Existing partners (matched by name) will be <b>updated</b>; new names will
    be <b>created</b>. Unmapped metrics are left blank for manual scoring later.</div>""", unsafe_allow_html=True)

    # Show import results from previous run
    if st.session_state.get("_import_done"):
        res = st.session_state.pop("_import_done")
        st.markdown(f'<div class="toast">✅ Import complete — {res["created"]} created, {res["updated"]} updated, {res["errors"]} errors</div>', unsafe_allow_html=True)

    uploaded = st.file_uploader("📁 Upload CSV file", type=["csv"], key="import_csv")
    if uploaded is None:
        st.info("Upload a CSV to get started. Required column: **Partner** (name or ID)."); st.stop()

    # Parse CSV
    try:
        df = pd.read_csv(uploaded)
    except Exception as e:
        st.error(f"Could not read CSV: {e}"); st.stop()
    if df.empty:
        st.error("The uploaded CSV is empty."); st.stop()

    csv_cols = list(df.columns)
    st.markdown("### 📄 CSV Preview (first 5 rows)")
    st.dataframe(df.head(5), use_container_width=True, hide_index=True)
    st.caption(f"{len(df)} rows × {len(csv_cols)} columns")

    # ── Mapping interface ──
    st.markdown("---")
    st.markdown("### 🔗 Column Mapping")
    none_opt = ["— None —"]

    # Helper: find best auto-match for a label among CSV columns
    def _auto_match(label, cols):
        ll = label.lower()
        for c in cols:
            if c.lower().strip() == ll: return c
        for c in cols:
            if ll in c.lower() or c.lower() in ll: return c
        return None

    # ── Required: Partner name column ──
    st.markdown('<div class="sec-head">🏢 Required</div>', unsafe_allow_html=True)
    auto_pn = _auto_match("partner", csv_cols) or _auto_match("partner name", csv_cols) or _auto_match("company", csv_cols)
    partner_col = st.selectbox("Partner name column **(required)**",
        csv_cols, index=csv_cols.index(auto_pn) if auto_pn else 0, key="imp_partner_col")

    # ── Optional: partner detail fields ──
    st.markdown('<div class="sec-head">📇 Partner Details (optional)</div>', unsafe_allow_html=True)
    detail_fields = [
        ("Year became partner", "partner_year", ["year","since","partner year","start year"]),
        ("Tier", "partner_tier", ["tier","level","designation"]),
        ("Partner Discount", "partner_discount", ["discount","partner discount","margin"]),
        ("City", "partner_city", ["city"]),
        ("Country", "partner_country", ["country"]),
        ("PAM name", "pam_name", ["pam","pam name","account manager","partner manager"]),
        ("PAM email", "pam_email", ["pam email","manager email"]),
    ]
    dc1, dc2, dc3 = st.columns(3)
    detail_mapping = {}
    for i, (label, field, hints) in enumerate(detail_fields):
        auto = None
        for h in hints:
            auto = _auto_match(h, csv_cols)
            if auto: break
        col_widget = [dc1, dc2, dc3][i % 3]
        with col_widget:
            sel = st.selectbox(label, none_opt + csv_cols,
                index=(csv_cols.index(auto) + 1) if auto and auto in csv_cols else 0,
                key=f"imp_det_{field}")
        if sel != "— None —":
            detail_mapping[field] = sel

    # ── Metric mapping ──
    st.markdown('<div class="sec-head">📊 Scoring Metrics</div>', unsafe_allow_html=True)
    st.caption("Map CSV columns to each metric. For quantitative metrics, the raw value from CSV will be auto-scored using your Step 1 ranges.")
    metric_mapping = {}
    mc1, mc2 = st.columns(2)
    for idx, m in enumerate(em):
        auto = _auto_match(m["name"], csv_cols)
        col_widget = mc1 if idx % 2 == 0 else mc2
        with col_widget:
            mtype = "📏" if m["type"] == "quantitative" else "📝"
            sel = st.selectbox(f'{mtype} {m["name"]}', none_opt + csv_cols,
                index=(csv_cols.index(auto) + 1) if auto and auto in csv_cols else 0,
                key=f"imp_m_{m['key']}")
            if sel == "— None —":
                st.markdown('<div style="height:3px;background:#DC2626;border-radius:2px;margin:-8px 0 6px;"></div>', unsafe_allow_html=True)
        if sel != "— None —":
            metric_mapping[m["key"]] = sel

    # Summary — highlight unmapped fields
    st.markdown("---")
    mapped_count = len(metric_mapping)
    unmapped_metrics = [m["name"] for m in em if m["key"] not in metric_mapping]
    unmapped_details = [label for (label, field, _) in detail_fields if field not in detail_mapping]
    st.markdown(f"**Mapped:** {mapped_count}/{len(em)} metrics  •  Partner column: **{partner_col}**")
    if unmapped_metrics:
        um_list = ", ".join(f"<b>{n}</b>" for n in unmapped_metrics)
        st.markdown(f'<div style="background:#FFF3CD;border-left:4px solid #FFA500;padding:12px 16px;border-radius:6px;margin:8px 0;font-size:.92rem;">'
                    f'⚠️ <b>Unmapped metrics ({len(unmapped_metrics)}):</b> {um_list}<br>'
                    f'<span style="color:#6a5a00;font-size:.84rem;">These will be left blank — you can score them manually later.</span></div>', unsafe_allow_html=True)
    if unmapped_details:
        ud_list = ", ".join(f"<b>{n}</b>" for n in unmapped_details)
        st.markdown(f'<div style="background:#E8F0FE;border-left:4px solid #4285F4;padding:12px 16px;border-radius:6px;margin:8px 0;font-size:.92rem;">'
                    f'ℹ️ <b>Unmapped detail fields ({len(unmapped_details)}):</b> {ud_list}</div>', unsafe_allow_html=True)
    if mapped_count == 0:
        st.markdown('<div style="background:#FDEDED;border-left:4px solid #D93025;padding:12px 16px;border-radius:6px;margin:8px 0;font-size:.92rem;">'
                    '🔴 <b>No metrics mapped</b> — partners will be created with blank scores for manual entry.</div>', unsafe_allow_html=True)
    elif not unmapped_metrics and not unmapped_details:
        st.markdown('<div style="background:#E6F4EA;border-left:4px solid #34A853;padding:12px 16px;border-radius:6px;margin:8px 0;font-size:.92rem;">'
                    '✅ <b>All fields mapped!</b> Ready to import.</div>', unsafe_allow_html=True)

    # ── Process import ──
    st.markdown("---")
    _, _, btn_col = st.columns([2, 2, 1])
    with btn_col:
        do_import = st.button("📥  Import Partners", use_container_width=True, type="primary")

    if do_import:
        created = 0; updated = 0; error_rows = []; skipped_limit = 0
        existing_names = {p.get("partner_name","").strip().lower() for p in _load_partners()}
        max_p = _max_partners(); current_count = len(existing_names)
        progress = st.progress(0, text="Importing...")

        for row_idx, row in df.iterrows():
            progress.progress(min((row_idx + 1) / len(df), 1.0), text=f"Processing row {row_idx + 1}/{len(df)}...")
            pname = str(row.get(partner_col, "")).strip()
            if not pname or pname.lower() in ("nan", "none", ""):
                error_rows.append({"row": row_idx + 2, "partner": pname, "error": "Missing partner name"})
                continue

            # Check partner limit for new partners (updates are always allowed)
            is_new = pname.strip().lower() not in existing_names
            if is_new and max_p and (current_count + created) >= max_p:
                skipped_limit += 1
                error_rows.append({"row": row_idx + 2, "partner": pname, "error": f"Partner limit ({max_p}) reached"})
                continue

            # Build raw dict
            raw_dict = {"partner_name": pname}
            for field, csv_col in detail_mapping.items():
                raw_dict[field] = str(row.get(csv_col, "")).strip()
                if raw_dict[field].lower() == "nan": raw_dict[field] = ""

            # Build scored dict
            row_dict = {"partner_name": pname}
            for field in ["partner_year","partner_tier","partner_discount","partner_city","partner_country","pam_name","pam_email"]:
                row_dict[field] = raw_dict.get(field, "")

            scores = {}; raw_vals = {}
            for m in em:
                mk = m["key"]
                if mk in metric_mapping:
                    raw_val = row.get(metric_mapping[mk])
                    if pd.isna(raw_val) or str(raw_val).strip().lower() in ("", "nan", "none", "n/a"):
                        raw_vals[mk] = None; scores[mk] = None
                    else:
                        raw_str = str(raw_val).strip()
                        raw_vals[mk] = raw_str
                        raw_dict[f"raw_{mk}"] = raw_str
                        scr = calc_score(mk, raw_str, cr)
                        scores[mk] = scr
                        row_dict[mk] = scr if scr else ""
                else:
                    raw_vals[mk] = None; scores[mk] = None
                    row_dict[mk] = ""

            # Calculate totals
            si = {k: v for k, v in scores.items() if v is not None}
            total = sum(si.values()); sn = len(si); mp = sn * 5
            pct = round(total / mp * 100, 1) if mp else 0
            row_dict["total_score"] = total
            row_dict["max_possible"] = mp
            row_dict["percentage"] = pct

            # Upsert
            try:
                is_update = not is_new
                _upsert_partner(row_dict, raw_dict)
                if is_update:
                    updated += 1
                else:
                    created += 1
                    existing_names.add(pname.strip().lower())
            except Exception as e:
                error_rows.append({"row": row_idx + 2, "partner": pname, "error": str(e)})

        progress.empty()

        # Show results
        st.session_state["_import_done"] = {"created": created, "updated": updated, "errors": len(error_rows)}

        st.markdown("### ✅ Import Results")
        r1, r2, r3 = st.columns(3)
        with r1: st.metric("Created", created)
        with r2: st.metric("Updated", updated)
        with r3: st.metric("Errors", len(error_rows))

        if skipped_limit > 0:
            st.warning(f"⚠️ **{skipped_limit}** partner(s) skipped — partner limit of **{max_p}** reached. Contact your admin to increase the limit.")

        if error_rows:
            st.markdown("#### ⚠️ Errors")
            err_df = pd.DataFrame(error_rows)
            st.dataframe(err_df, use_container_width=True, hide_index=True)
            st.download_button("⬇️ Download Errors CSV", err_df.to_csv(index=False),
                "import_errors.csv", "text/csv")

        if created > 0 or updated > 0:
            st.success(f"Successfully imported **{created + updated}** partners. View them on the **Partner List** page or in **Step 3**.")

            # Show imported data with actual values (not just scores)
            st.markdown("#### 📊 Imported Data — Actual Values")
            st.caption("Shows the raw values imported from your CSV alongside the computed scores (1–5).")
            imported_partners = _load_partners()
            raw_all = _load_raw()
            preview_rows = []
            for p in imported_partners:
                pn = p.get("partner_name","")
                raw_p = next((r for r in raw_all if r.get("partner_name","") == pn), {})
                row = {"Partner": pn}
                for m in em:
                    raw_key = f"raw_{m['key']}"
                    raw_val = raw_p.get(raw_key, "")
                    try: score_val = int(p.get(m["key"],"") or 0)
                    except: score_val = 0
                    if raw_val:
                        row[m["name"]] = f"{raw_val}  →  {score_val}/5" if score_val else str(raw_val)
                    else:
                        row[m["name"]] = f"{score_val}/5" if score_val else "—"
                try: pct = float(p.get("percentage",0) or 0)
                except: pct = 0
                row["Total"] = p.get("total_score", 0)
                row["Pct"] = f"{pct:.1f}%"
                preview_rows.append(row)
            if preview_rows:
                preview_df = pd.DataFrame(preview_rows)
                st.dataframe(preview_df, use_container_width=True, hide_index=True)

    # ── Recalculate Benchmarks ──
    if _raw_path().exists():
        st.markdown("---")
        st.markdown("### 📐 Recalculate Scoring Benchmarks")
        st.markdown("""<div class="info-box">
        Analyze the distribution of your imported partner data and automatically set
        <b>quintile-based scoring ranges</b> (1–5) for all quantitative metrics.
        This replaces the manual thresholds in Step 1 with data-driven boundaries
        where each score bucket contains roughly 20% of your partners.<br><br>
        <b>Qualitative</b> metrics (descriptors) are not affected.
        </div>""", unsafe_allow_html=True)

        raw_all = _load_raw()
        st.caption(f"Partner data available: **{len(raw_all)}** partners")

        if st.session_state.get("_bench_result"):
            res = st.session_state.pop("_bench_result")
            if res["updated"]:
                st.markdown(f'<div class="toast">✅ Benchmarks recalculated — **{len(res["updated"])}** metric(s) updated, all partners re-scored</div>', unsafe_allow_html=True)
                with st.expander("Updated metrics", expanded=False):
                    for name in res["updated"]:
                        st.markdown(f"- {name}")
            else:
                st.info("No quantitative metrics were changed (not enough data variation or no mapped metrics).")
            if res["skipped"]:
                with st.expander(f"Unchanged metrics ({len(res['skipped'])})", expanded=False):
                    for name in res["skipped"]:
                        st.markdown(f"- {name}")

        _, _, bench_col = st.columns([2, 2, 1])
        with bench_col:
            do_bench = st.button("📐  Recalculate Benchmarks", use_container_width=True, type="secondary")
        if do_bench:
            with st.spinner("Analyzing partner data distributions..."):
                result = _recalculate_benchmarks()
            st.session_state["_bench_result"] = result
            st.rerun()


# ═════════════════════════════════════════════════════════════════════════
# PARTNER LIST — sortable table + inline edit + manual add
# ═════════════════════════════════════════════════════════════════════════
elif page=="Partner List":
    _brand(); st.markdown("## Partner List")
    if not _save_path().exists():
        st.warning("⚠️ Complete **Step 1 — Scoring Criteria** first."); st.stop()
    st.session_state["criteria"] = json.loads(_save_path().read_text())
    _ensure_criteria_complete()
    cr = st.session_state["criteria"]; em = _enabled()
    partners = _load_partners()

    # ── Edit mode (single partner) ──
    edit_pn = st.session_state.get("_pl_edit")
    if edit_pn:
        raw_all = _load_raw()
        raw_p = next((r for r in raw_all if r.get("partner_name") == edit_pn), None)
        csv_p = next((p for p in partners if p.get("partner_name") == edit_pn), None)
        if not csv_p:
            st.error(f"Partner '{edit_pn}' not found."); st.session_state.pop("_pl_edit", None); st.rerun()

        if st.button("← Back to Partner List", key="pl_back"):
            st.session_state.pop("_pl_edit", None); st.rerun()

        st.markdown(f"### ✏️ Edit Scorecard — {edit_pn}")

        # Show save confirmation
        if st.session_state.get("_pl_edit_saved"):
            st.markdown('<div class="toast">✅ Partner scorecard updated</div>', unsafe_allow_html=True)
            st.session_state["_pl_edit_saved"] = False

        with st.form("pl_edit_form"):
            # Partner details
            st.markdown('<div class="sec-head">📇 Partner Details</div>', unsafe_allow_html=True)
            d1, d2, d3 = st.columns(3)
            with d1:
                ed_name = st.text_input("Partner name", value=edit_pn, key="ple_name", disabled=True)
                ed_year = st.text_input("Year became partner", value=csv_p.get("partner_year",""), key="ple_year")
            with d2:
                tiers = _tiers(); t_opts = [""] + tiers if tiers else [""]
                cur_tier = csv_p.get("partner_tier","")
                ed_tier = st.selectbox("Tier", t_opts, index=t_opts.index(cur_tier) if cur_tier in t_opts else 0, key="ple_tier")
                ed_disc = st.text_input("Partner Discount", value=csv_p.get("partner_discount",""), key="ple_disc")
                ed_city = st.text_input("City", value=csv_p.get("partner_city",""), key="ple_city")
            with d3:
                cur_country = csv_p.get("partner_country","")
                ed_country = st.selectbox("Country", COUNTRIES, index=COUNTRIES.index(cur_country) if cur_country in COUNTRIES else 0, key="ple_country")
                ed_pam = st.text_input("PAM name", value=csv_p.get("pam_name",""), key="ple_pam")
            ed_pam_email = st.text_input("PAM email", value=csv_p.get("pam_email",""), key="ple_pam_email")

            # Metric scores — show actual values (raw) with computed scores
            st.markdown('<div class="sec-head">📊 Metric Scores — enter the actual value; the score (1–5) is computed automatically</div>', unsafe_allow_html=True)
            edit_raw_vals = {}
            for cat in CATEGORIES:
                cat_metrics = [m for m in em if m["key"] in cat["keys"]]
                if not cat_metrics: continue
                st.markdown(f"**{cat['icon']} {cat['label']}**")
                for m in cat_metrics:
                    mk = m["key"]; mc = cr.get(mk, {})
                    if not mc: continue
                    iq = m["type"] == "quantitative"
                    # Get the raw value from raw JSON (the actual imported value)
                    raw_val = raw_p.get(f"raw_{mk}", "") if raw_p else ""
                    # Fallback: if no raw value stored, try to reverse from score
                    if not raw_val and csv_p:
                        try:
                            cur_score = int(csv_p.get(mk, 0) or 0)
                            if cur_score and 1 <= cur_score <= 5:
                                raw_val = str(_synthetic_raw_for_score(mk, cur_score, cr) or "")
                        except: pass

                    if iq:
                        u = mc.get("unit", "") or ""
                        hints = []
                        for s in ("1","2","3","4","5"):
                            r = mc["ranges"][s]; lo, hi = r["min"], r["max"]
                            if lo and hi: hints.append(f"**{s}**: {lo}–{hi}")
                            elif lo and not hi: hints.append(f"**{s}**: ≥{lo}")
                            elif not lo and hi: hints.append(f"**{s}**: ≤{hi}")
                        ic, sc_c = st.columns([4, 1])
                        with ic:
                            pv = st.text_input(f"{m['name']} ({u})" if u else m["name"],
                                value=str(raw_val) if raw_val else "",
                                key=f"ple_v_{mk}",
                                help=(f"{m['explanation']}  •  Ranges: {', '.join(hints)}" if hints else m["explanation"]) if is_admin else m["explanation"])
                        scr = calc_score(mk, pv, cr) if pv else None
                        with sc_c:
                            if scr:
                                st.markdown(f'<div style="margin-top:28px;text-align:center;padding:6px 0;border-radius:6px;font-weight:800;color:#fff;background:{SC.get(scr,"#ccc")}">{scr}/5</div>', unsafe_allow_html=True)
                            else:
                                st.markdown('<div style="margin-top:28px;text-align:center;padding:6px 0;border-radius:6px;font-weight:800;color:#888;background:#eee">—</div>', unsafe_allow_html=True)
                        edit_raw_vals[mk] = pv
                    else:
                        opts = ["— Select —"] + [f"({s}) {mc['descriptors'][s]}" for s in ("1","2","3","4","5")]
                        # Pre-select matching descriptor
                        sel_idx = 0
                        if raw_val:
                            for oi, o in enumerate(opts):
                                if oi == 0: continue
                                desc = re.sub(r"^\(\d\)\s*", "", o)
                                if desc == raw_val:
                                    sel_idx = oi; break
                            else:
                                for oi, o in enumerate(opts):
                                    if oi > 0 and raw_val in o: sel_idx = oi; break
                        ic, sc_c = st.columns([4, 1])
                        with ic:
                            pv = st.selectbox(m["name"], opts, index=sel_idx,
                                key=f"ple_v_{mk}", help=m["explanation"])
                        if pv and pv != "— Select —":
                            raw_d = re.sub(r"^\(\d\)\s*", "", pv)
                            scr = calc_score(mk, raw_d, cr)
                        else:
                            raw_d = None; scr = None
                        with sc_c:
                            if scr:
                                st.markdown(f'<div style="margin-top:28px;text-align:center;padding:6px 0;border-radius:6px;font-weight:800;color:#fff;background:{SC.get(scr,"#ccc")}">{scr}/5</div>', unsafe_allow_html=True)
                            else:
                                st.markdown('<div style="margin-top:28px;text-align:center;padding:6px 0;border-radius:6px;font-weight:800;color:#888;background:#eee">—</div>', unsafe_allow_html=True)
                        edit_raw_vals[mk] = raw_d

            st.markdown("---")
            _, save_col = st.columns([3, 1])
            with save_col:
                edit_sub = st.form_submit_button("💾  Save Changes", use_container_width=True, type="primary")

        if edit_sub:
            # Build updated raw dict and row dict from actual values
            new_raw = {"partner_name": edit_pn, "partner_year": ed_year, "partner_tier": ed_tier,
                       "partner_discount": ed_disc,
                       "partner_city": ed_city, "partner_country": ed_country,
                       "pam_name": ed_pam, "pam_email": ed_pam_email}
            new_row = dict(new_raw)  # start with same detail fields
            si = {}
            for m in em:
                mk = m["key"]
                raw_val = st.session_state.get(f"ple_v_{mk}", "")
                # Handle qualitative selectbox values
                if isinstance(raw_val, str) and raw_val.startswith("(") and m["type"] == "qualitative":
                    raw_val = re.sub(r"^\(\d\)\s*", "", raw_val)
                if raw_val and raw_val != "— Select —":
                    new_raw[f"raw_{mk}"] = raw_val
                    scr = calc_score(mk, raw_val, cr)
                    if scr and 1 <= scr <= 5:
                        new_row[mk] = scr; si[mk] = scr
                    else:
                        new_row[mk] = ""
                else:
                    new_raw[f"raw_{mk}"] = None
                    new_row[mk] = ""
            total = sum(si.values()); sn = len(si); mp = sn * 5
            new_row["total_score"] = total; new_row["max_possible"] = mp
            new_row["percentage"] = round(total / mp * 100, 1) if mp else 0
            _upsert_partner(new_row, new_raw)
            st.session_state["_pl_edit_saved"] = True; st.rerun()

    else:
        # ── Partner table view ──
        if st.session_state.get("_pl_added"):
            st.markdown('<div class="toast">✅ New partner created</div>', unsafe_allow_html=True)
            st.session_state["_pl_added"] = False

        if not partners:
            st.info("No partners yet. Use **Import Data** to bulk-import or add one manually below.")
        else:
            # Build display dataframe
            tbl_data = []
            for p in sorted(partners, key=lambda x: -int(x.get("total_score", 0) or 0)):
                try: pct = float(p.get("percentage", 0) or 0)
                except: pct = 0
                gl, gc = _grade(pct)
                try: ts = int(p.get("total_score", 0) or 0)
                except: ts = 0
                tbl_data.append({
                    "Partner": p.get("partner_name",""),
                    "Tier": p.get("partner_tier",""),
                    "PAM": p.get("pam_name",""),
                    "City": p.get("partner_city",""),
                    "Total Score": ts,
                    "Percentage": f"{pct:.1f}%",
                    "Grade": gl,
                })

            st.dataframe(pd.DataFrame(tbl_data), use_container_width=True, hide_index=True,
                column_config={
                    "Partner": st.column_config.TextColumn(width="large"),
                    "Total Score": st.column_config.NumberColumn(format="%d"),
                })

            # Edit / Delete controls
            st.markdown("---")
            st.markdown("### ✏️ Edit or Delete a Partner")
            partner_names = sorted([p.get("partner_name","") for p in partners])
            sel_partner = st.selectbox("Select partner", partner_names, key="pl_sel_partner")
            ec1, ec2 = st.columns(2)
            with ec1:
                if st.button("✏️  Edit Scorecard", use_container_width=True, type="primary", key="pl_edit_btn"):
                    st.session_state["_pl_edit"] = sel_partner; st.rerun()
            with ec2:
                if st.button("🗑️  Delete Partner", use_container_width=True, key="pl_del_btn"):
                    _delete_partner(sel_partner); st.rerun()

        # ── Manual add ──
        st.markdown("---")
        st.markdown("### ➕ Add New Partner")
        with st.form("pl_add_form"):
            a1, a2, a3 = st.columns(3)
            with a1:
                new_name = st.text_input("Partner name *", key="pla_name", placeholder="e.g. Acme Corp")
            with a2:
                new_year = st.text_input("Year became partner", key="pla_year", placeholder="e.g. 2022")
            with a3:
                tiers = _tiers(); t_opts = ["Please choose..."] + tiers if tiers else ["Please choose..."]
                new_tier = st.selectbox("Tier", t_opts, key="pla_tier")
            a4, a5, a6 = st.columns(3)
            with a4: new_disc = st.text_input("Partner Discount", key="pla_disc", placeholder="e.g. 20%")
            with a5: new_city = st.text_input("City", key="pla_city")
            with a6:
                new_country = st.selectbox("Country", COUNTRIES, key="pla_country")
            a7, a8 = st.columns(2)
            with a7: new_pam = st.text_input("PAM name", key="pla_pam")

            _, add_col = st.columns([3, 1])
            with add_col:
                add_sub = st.form_submit_button("➕  Add Partner", use_container_width=True, type="primary")

        if add_sub:
            if not new_name or not new_name.strip():
                st.error("Partner name is required.")
            elif _partner_exists(new_name):
                st.error(f"A partner named **{new_name}** already exists.")
            elif _max_partners() and _partner_count() >= _max_partners():
                st.error(f"Partner limit reached (**{_max_partners()}**). Contact your admin to increase the limit.")
            else:
                row = {"partner_name": new_name.strip(), "partner_year": new_year,
                       "partner_tier": new_tier if new_tier != "Please choose..." else "",
                       "partner_discount": new_disc,
                       "partner_city": new_city, "partner_country": new_country,
                       "pam_name": new_pam, "pam_email": "",
                       "total_score": 0, "max_possible": 0, "percentage": 0}
                for m in em: row[m["key"]] = ""
                raw = {"partner_name": new_name.strip(), "partner_year": new_year,
                       "partner_tier": new_tier if new_tier != "Please choose..." else "",
                       "partner_discount": new_disc,
                       "partner_city": new_city, "partner_country": new_country,
                       "pam_name": new_pam, "pam_email": ""}
                _append_partner(row, raw)
                st.session_state["_pl_added"] = True; st.rerun()


# ═════════════════════════════════════════════════════════════════════════
# ASK CHANNELPRO — AI-powered natural language data agent
# ═════════════════════════════════════════════════════════════════════════
elif page=="Ask ChannelPRO™":
    _brand(); st.markdown("## 🤖 Ask ChannelPRO™")
    if not _save_path().exists():
        st.warning("⚠️ Complete **Step 1 — Scoring Criteria** first."); st.stop()
    st.session_state["criteria"] = json.loads(_save_path().read_text())
    _ensure_criteria_complete()
    cr = st.session_state["criteria"]
    partners = _load_partners()
    if not partners:
        st.warning("⚠️ Score at least one partner in **Step 2** before using the AI assistant."); st.stop()

    st.markdown("""<div class="info-box">
    Ask questions about your partner scorecards in plain English. Examples:<br>
    • <i>"Which partners have MDF utilization below 40%?"</i><br>
    • <i>"Show me partners with both a low close rate and long sales cycle"</i><br>
    • <i>"Compare the top 5 partners by revenue vs their customer satisfaction"</i><br>
    • <i>"Set Partner X's renewal rate score to 4"</i><br>
    Conversations are multi-turn — ask follow-up questions to refine results.</div>""", unsafe_allow_html=True)

    # API key management — env var → session cache → user input (see utils.api)
    api_key = resolve_api_key()

    # Init chat history
    if "ai_messages" not in st.session_state:
        st.session_state["ai_messages"] = []
    if "ai_pending_updates" not in st.session_state:
        st.session_state["ai_pending_updates"] = None

    # ── Pending updates confirmation ──
    if st.session_state["ai_pending_updates"]:
        updates = st.session_state["ai_pending_updates"]
        st.markdown("### ⚠️ Confirm Score Updates")
        st.markdown("The AI has suggested the following changes:")
        upd_rows = ""
        for u in updates:
            upd_rows += f'<tr><td style="text-align:left;padding-left:10px">{u["partner"]}</td>'
            mk = u["metric_key"]
            mname = next((m["name"] for m in SCORECARD_METRICS if m["key"] == mk), mk)
            upd_rows += f'<td>{mname}</td><td style="font-weight:800">{u["new_score"]}/5</td>'
            upd_rows += f'<td style="text-align:left">{u.get("reason","")}</td></tr>'
        st.markdown(f'<table class="hm-tbl"><thead><tr><th style="text-align:left">Partner</th><th>Metric</th><th>New Score</th><th style="text-align:left">Reason</th></tr></thead><tbody>{upd_rows}</tbody></table>', unsafe_allow_html=True)
        uc1, uc2, uc3 = st.columns([1, 1, 3])
        with uc1:
            if st.button("✅ Apply Updates", type="primary", key="ai_confirm_upd"):
                applied = _apply_ai_updates(updates, cr)
                st.session_state["ai_pending_updates"] = None
                st.session_state["ai_messages"].append({"role": "assistant", "content":
                    json.dumps({"answer": f"✅ Applied {applied} score update(s) successfully.", "table": None, "chart": None, "updates": None})})
                st.rerun()
        with uc2:
            if st.button("❌ Cancel", key="ai_cancel_upd"):
                st.session_state["ai_pending_updates"] = None
                st.session_state["ai_messages"].append({"role": "assistant", "content":
                    json.dumps({"answer": "Updates cancelled. No changes were made.", "table": None, "chart": None, "updates": None})})
                st.rerun()
        st.markdown("---")

    # ── Chat history display ──
    for msg in st.session_state["ai_messages"]:
        if msg["role"] == "user":
            with st.chat_message("user"):
                st.markdown(msg["content"])
        else:
            with st.chat_message("assistant", avatar="🤖"):
                try:
                    resp = json.loads(msg["content"])
                except (json.JSONDecodeError, TypeError):
                    resp = None
                if resp and isinstance(resp, dict):
                    answer_text = resp.get("answer", "")
                    if answer_text:
                        st.markdown(answer_text.replace("\\n", "\n"))
                    if resp.get("table"):
                        try:
                            st.dataframe(pd.DataFrame(resp["table"]), use_container_width=True, hide_index=True)
                        except Exception:
                            pass
                    if resp.get("chart"):
                        try:
                            _render_ai_chart(resp["chart"])
                        except Exception:
                            pass
                else:
                    # Plain text message (non-JSON)
                    st.markdown(msg["content"])

    # ── Controls row ──
    ctrl1, ctrl2 = st.columns([1, 6])
    with ctrl1:
        if st.button("🗑️ Clear chat", key="ai_clear", use_container_width=True):
            st.session_state["ai_messages"] = []
            st.session_state["ai_pending_updates"] = None
            st.rerun()

    # ── Chat input ──
    user_input = st.chat_input("Ask about your partners...", key="ai_chat_input")
    if user_input:
        # Display user message inline
        with st.chat_message("user"):
            st.markdown(user_input)
        # Add user message to history
        st.session_state["ai_messages"].append({"role": "user", "content": user_input})

        # Build messages for API (only role + content for API call)
        api_messages = []
        for msg in st.session_state["ai_messages"]:
            if msg["role"] == "user":
                api_messages.append({"role": "user", "content": msg["content"]})
            else:
                # Send back the raw JSON as assistant response for context
                api_messages.append({"role": "assistant", "content": msg["content"]})

        # Call API and display response inline
        with st.chat_message("assistant", avatar="🤖"):
            with st.spinner("🤖 Analyzing your partner data..."):
                resp = _call_ai(api_messages, api_key)
            st.markdown(resp.get("answer","").replace("\\n", "\n"))
            if resp.get("table"):
                try:
                    st.dataframe(pd.DataFrame(resp["table"]), use_container_width=True, hide_index=True)
                except Exception:
                    pass
            if resp.get("chart"):
                try:
                    _render_ai_chart(resp["chart"])
                except Exception:
                    pass

        # Store response
        resp_json = json.dumps(resp)
        st.session_state["ai_messages"].append({"role": "assistant", "content": resp_json})

        # Handle updates — stage for confirmation (requires rerun to show confirmation UI)
        if resp.get("updates") and isinstance(resp["updates"], list) and len(resp["updates"]) > 0:
            st.session_state["ai_pending_updates"] = resp["updates"]
            st.rerun()


# ═════════════════════════════════════════════════════════════════════════
# ADMIN — MANAGE USERS
# ═════════════════════════════════════════════════════════════════════════
elif page=="Admin — Manage Users":
    _brand(); st.markdown("## Admin — Manage Users & Clients")
    if not is_admin: st.error("Admin access required."); st.stop()
    users=_load_users()
    if st.session_state.get("_admin_saved"):
        st.markdown('<div class="toast">✅ Changes saved</div>',unsafe_allow_html=True); st.session_state["_admin_saved"]=False
    st.markdown("### Current Users")
    for uname,udata in users.items():
        with st.expander(f"{'🔑' if udata['role']=='admin' else '👤'} **{uname}** — {udata['display_name']} ({udata['role']}) {('→ '+udata['tenant']) if udata.get('tenant') else ''}"):
            if uname=="admin": st.caption("Default admin account. You can change the password below.")
            new_pw=st.text_input(f"New password for {uname}",type="password",key=f"adm_pw_{uname}")
            if st.button(f"Update password",key=f"adm_pwbtn_{uname}"):
                if new_pw and len(new_pw)>=4:
                    users[uname]["password_hash"]=_hash_pw(new_pw); _save_users(users)
                    st.session_state["_admin_saved"]=True; st.rerun()
                else: st.error("Password must be at least 4 characters.")
            if uname!="admin":
                if st.button(f"🗑️ Delete user {uname}",key=f"adm_del_{uname}"):
                    del users[uname]; _save_users(users)
                    st.session_state["_admin_saved"]=True; st.rerun()
    st.markdown("---")
    st.markdown("### Add New Client User")
    st.markdown("Each client user gets a **tenant ID** (e.g. `acme_corp`). This isolates their data completely.")
    with st.form("add_user_form"):
        c1,c2=st.columns(2)
        with c1:
            new_uname=st.text_input("Username",placeholder="e.g. acme_user")
            new_display=st.text_input("Display name",placeholder="e.g. Acme Corporation")
        with c2:
            new_password=st.text_input("Password",type="password",placeholder="Min 4 characters")
            new_tenant=st.text_input("Tenant ID",placeholder="e.g. acme_corp (lowercase, no spaces)")
        new_role=st.radio("Role",["client","admin"],horizontal=True)
        add_sub=st.form_submit_button("➕ Add User",type="primary")
    if add_sub:
        errors=[]
        if not new_uname or len(new_uname)<2: errors.append("Username must be at least 2 characters.")
        if new_uname in users: errors.append("Username already exists.")
        if not new_password or len(new_password)<4: errors.append("Password must be at least 4 characters.")
        if new_role=="client" and (not new_tenant or " " in new_tenant): errors.append("Tenant ID is required for client users (no spaces).")
        if new_uname and not re.match(r'^[a-zA-Z0-9_]+$', new_uname): errors.append("Username: letters, numbers, underscores only.")
        if errors:
            for e in errors: st.error(e)
        else:
            tid=new_tenant.lower().strip() if new_tenant else None
            users[new_uname]={"password_hash":_hash_pw(new_password),"display_name":new_display or new_uname,"role":new_role,"tenant":tid if new_role=="client" else None}
            _save_users(users)
            if tid: _tenant_dir(tid)
            st.session_state["_admin_saved"]=True; st.rerun()
    st.markdown("---")
    st.markdown("### Tenant Directories & Limits")
    tenants=_all_tenants()
    if tenants:
        for t in tenants:
            td=_tenant_dir(t)
            has_criteria=(td/"scoring_criteria.json").exists()
            has_partners=(td/"all_partners.csv").exists()
            pc=len(_load_partners(td/"all_partners.csv")) if has_partners else 0
            tcfg = _load_tenant_config(t)
            mp = tcfg.get("max_partners", 0)
            limit_str = f"**{mp}**" if mp else "Unlimited"
            with st.expander(f"**{t}** — {'✅ Criteria' if has_criteria else '⚪ No criteria'} · {pc} partners · Limit: {limit_str}"):
                new_max = st.number_input(
                    f"Max partners for **{t}** (0 = unlimited)",
                    min_value=0, max_value=10000, value=mp, step=5, key=f"adm_maxp_{t}",
                    help="Set the maximum number of partners this client can score. 0 means unlimited.")
                if st.button(f"💾 Save limit for {t}", key=f"adm_maxp_btn_{t}"):
                    tcfg["max_partners"] = new_max
                    _save_tenant_config(tcfg, t)
                    st.session_state["_admin_saved"] = True; st.rerun()
    else:
        st.info("No tenant directories yet. Add a client user above to create one.")

# ═════════════════════════════════════════════════════════════════════════
# ADMIN — ALL CLIENTS OVERVIEW
# ═════════════════════════════════════════════════════════════════════════
elif page=="Admin — All Clients":
    _brand(); st.markdown("## Admin — All Clients Overview")
    if not is_admin: st.error("Admin access required."); st.stop()
    tenants=_all_tenants()
    if not tenants: st.info("No clients yet. Create accounts in **Manage Users**."); st.stop()
    total_partners=0; tenant_data={}
    for t in tenants:
        td=_tenant_dir(t)
        ci=json.loads((td/"client_info.json").read_text()) if (td/"client_info.json").exists() else {}
        ps=_load_partners(td/"all_partners.csv")
        total_partners+=len(ps)
        # Load break-even data if available
        be_file = td / "break_even_configs.json"
        be_data = json.loads(be_file.read_text()) if be_file.exists() else None
        be_total = sum(sum(v for v in items.values()) for items in be_data.get("sections", {}).values()) if be_data else 0
        be_np = be_data.get("num_partners", 0) if be_data else 0
        tenant_data[t]={"client_info":ci,"partners":ps,"has_criteria":(td/"scoring_criteria.json").exists(),
                        "be_data":be_data,"be_total":be_total,"be_np":be_np}
    c1,c2,c3=st.columns(3)
    with c1: st.markdown(f'<div class="sum-card"><div class="sum-big">{len(tenants)}</div><div class="sum-lbl">Clients</div></div>',unsafe_allow_html=True)
    with c2: st.markdown(f'<div class="sum-card"><div class="sum-big">{total_partners}</div><div class="sum-lbl">Total Partners</div></div>',unsafe_allow_html=True)
    with c3:
        wc=sum(1 for d in tenant_data.values() if d["has_criteria"])
        st.markdown(f'<div class="sum-card"><div class="sum-big">{wc}/{len(tenants)}</div><div class="sum-lbl">Criteria Set</div></div>',unsafe_allow_html=True)
    # Break-even overview row
    be_clients = sum(1 for d in tenant_data.values() if d["be_data"])
    if be_clients > 0:
        total_be = sum(d["be_total"] for d in tenant_data.values())
        b1, b2, b3 = st.columns(3)
        with b1: st.metric("Clients w/ Break-even", f"{be_clients}/{len(tenants)}")
        with b2: st.metric("Total Program Costs (all)", f"${total_be:,.0f}")
        avg_be = total_be / sum(d["be_np"] for d in tenant_data.values() if d["be_np"] > 0) if any(d["be_np"] > 0 for d in tenant_data.values()) else 0
        with b3: st.metric("Avg Break-even/Partner", f"${avg_be:,.2f}")
    st.markdown("---")
    for t in tenants:
        td=tenant_data[t]; ci=td["client_info"]; ps=td["partners"]
        client_name=ci.get("client_name",t)
        is_active = (t == active_tenant)
        active_tag = ' <span style="background:#1E293B;color:#fff;font-size:.72rem;font-weight:700;padding:2px 10px;border-radius:12px;margin-left:8px;vertical-align:middle;">ACTIVE</span>' if is_active else ""
        label_style = "color:#000;font-weight:800" if is_active else ""
        with st.expander(f"🏢 **{client_name}** ({t}) — {len(ps)} partners {'✅' if td['has_criteria'] else '⚪'}", expanded=is_active):
            if is_active:
                st.markdown(f'<div style="font-size:.88rem;{label_style};margin-bottom:8px;">Currently active client{active_tag}</div>', unsafe_allow_html=True)
            if ci:
                c1,c2,c3=st.columns(3)
                with c1: st.markdown(f"**Contact:** {ci.get('project_manager','—')}")
                with c2: st.markdown(f"**Email:** {ci.get('email','—')}")
                with c3: st.markdown(f"**City:** {ci.get('city','—')}, {ci.get('country','—')}")
            if ps:
                sorted_ps=sorted(ps,key=lambda p:-int(p.get("total_score",0) or 0))
                tbl="<table class='hm-tbl'><thead><tr><th>Partner</th><th>PAM</th><th>Total</th><th>%</th><th>Grade</th></tr></thead><tbody>"
                for p in sorted_ps:
                    try: tv=int(p.get("total_score",0) or 0)
                    except: tv=0
                    try: pv=float(p.get("percentage",0) or 0)
                    except: pv=0
                    gl,gc=_grade(pv)
                    tbl+=f'<tr><td style="text-align:left;padding-left:10px">{p.get("partner_name","")}</td><td>{p.get("pam_name","")}</td><td>{tv}</td><td>{pv:.1f}%</td><td style="color:{gc};font-weight:800">{gl}</td></tr>'
                tbl+="</tbody></table>"
                st.markdown(tbl,unsafe_allow_html=True)
                csv_path=_tenant_dir(t)/"all_partners.csv"
                if csv_path.exists(): st.download_button(f"⬇️ Download {t} CSV",csv_path.read_text(),f"{t}_partners.csv","text/csv",key=f"dl_{t}")
            else: st.caption("No partners scored yet.")
            # Break-even summary
            if td.get("be_data"):
                st.markdown("---")
                bc1, bc2, bc3 = st.columns(3)
                with bc1: st.metric("Program Costs", f"${td['be_total']:,.0f}")
                be_pt = td['be_total'] / td['be_np'] if td['be_np'] > 0 else 0
                with bc2: st.metric("Break-even/Partner", f"${be_pt:,.2f}")
                sup_t = sum(td["be_data"].get("sections",{}).get("Technical and Sales Support",{}).values())
                with bc3: st.metric("Support Costs", f"${sup_t:,.0f}")
    st.markdown("---"); st.markdown("### Cross-Client Export")
    if st.button("⬇️  Export All Clients to Single Excel",type="primary"):
        try:
            import openpyxl
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            wb=openpyxl.Workbook(); first=True
            for t in tenants:
                td_info=tenant_data[t]; ps=td_info["partners"]
                if not ps: continue
                cn=td_info["client_info"].get("client_name",t)[:31]
                if first: ws=wb.active; ws.title=cn; first=False
                else: ws=wb.create_sheet(title=cn)
                hf=PatternFill(start_color="1E2A3A",end_color="1E2A3A",fill_type="solid")
                hfont=Font(color="FFFFFF",bold=True,size=10)
                bdr=Border(left=Side(style="thin",color="CCCCCC"),right=Side(style="thin",color="CCCCCC"),top=Side(style="thin",color="CCCCCC"),bottom=Side(style="thin",color="CCCCCC"))
                headers=["Partner","PAM","Total Score","Percentage","Grade"]
                for ci_idx,h in enumerate(headers,1):
                    c=ws.cell(1,ci_idx,h); c.fill=hf; c.font=hfont; c.border=bdr; c.alignment=Alignment(horizontal="center")
                ws.column_dimensions["A"].width=28; ws.column_dimensions["B"].width=22
                sorted_ps=sorted(ps,key=lambda p:-int(p.get("total_score",0) or 0))
                for ri,p in enumerate(sorted_ps,2):
                    try: tv=int(p.get("total_score",0) or 0)
                    except: tv=0
                    try: pv=float(p.get("percentage",0) or 0)
                    except: pv=0
                    gl,_=_grade(pv)
                    ws.cell(ri,1,p.get("partner_name","")).border=bdr
                    ws.cell(ri,2,p.get("pam_name","")).border=bdr
                    ws.cell(ri,3,tv).border=bdr; ws.cell(ri,3).alignment=Alignment(horizontal="center")
                    pc=ws.cell(ri,4); pc.value=pv/100; pc.number_format="0.0%"; pc.border=bdr; pc.alignment=Alignment(horizontal="center")
                    ws.cell(ri,5,gl).border=bdr; ws.cell(ri,5).alignment=Alignment(horizontal="center")
            buf=io.BytesIO(); wb.save(buf)
            st.download_button("📥 Download",buf.getvalue(),"All_Clients_Overview.xlsx","application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        except ImportError: st.warning("Install openpyxl for Excel export.")


# ═════════════════════════════════════════════════════════════════════════
# BREAK-EVEN — PROGRAM COSTS
# ═════════════════════════════════════════════════════════════════════════
elif page=="Break-even — Program Costs":
    _brand(); st.markdown("## Break-even — Program Costs")
    st.markdown("""<div class="info-box">
    Enter your annual partner program costs by category. The tool calculates your <b>total program cost</b>
    and <b>nominal break-even point</b> (cost per partner). For <b>Technical and Sales Support</b>, it also
    derives cost-per-call and cost-per-minute metrics used in the Detailed Analysis page.</div>""", unsafe_allow_html=True)

    cfg = _load_be()
    if st.session_state.get("_be_saved"):
        st.markdown('<div class="toast">✅ Break-even configuration saved</div>', unsafe_allow_html=True)
        st.session_state["_be_saved"] = False

    # Custom categories in session state
    if "be_custom" not in st.session_state:
        st.session_state["be_custom"] = cfg.get("custom_items", {})

    # --- Add custom category UI (outside form) ---
    with st.expander("➕ Add custom cost category"):
        ac1, ac2 = st.columns(2)
        sec_names = [s["section"] for s in BE_SECTIONS] + ["— New section —"]
        with ac1: add_sec = st.selectbox("Section", sec_names, key="be_add_sec")
        with ac2:
            new_sec_name = ""
            if add_sec == "— New section —":
                new_sec_name = st.text_input("New section name", key="be_new_sec")
            add_item = st.text_input("Cost category name", key="be_add_item")
        if st.button("➕ Add", key="be_add_btn"):
            target = new_sec_name.strip() if add_sec == "— New section —" else add_sec
            if target and add_item.strip():
                custom = st.session_state["be_custom"]
                if target not in custom: custom[target] = {}
                custom[target][add_item.strip()] = 0
                st.rerun()

    # --- Main cost form ---
    with st.form("be_form"):
        new_cfg = {"sections": {}, "custom_items": st.session_state.get("be_custom", {})}

        grand_total = 0
        support_subtotal = 0

        # Iterate default sections + custom-only sections
        all_sec_names = [s["section"] for s in BE_SECTIONS]
        custom = st.session_state.get("be_custom", {})
        for csk in custom:
            if csk not in all_sec_names: all_sec_names.append(csk)

        for sec_name in all_sec_names:
            # Find default items
            default_sec = next((s for s in BE_SECTIONS if s["section"] == sec_name), None)
            default_items = default_sec["items"] if default_sec else []
            custom_items = list(custom.get(sec_name, {}).keys())
            all_items = default_items + [ci for ci in custom_items if ci not in default_items]

            if not all_items: continue

            icon = BE_SECTION_ICONS.get(sec_name, "📁")
            st.markdown(f'<div class="sec-head">{icon} {sec_name}</div>', unsafe_allow_html=True)

            sec_data = {}; sec_total = 0
            saved_sec = cfg.get("sections", {}).get(sec_name, {})
            # Layout: 2 columns of items
            cols = st.columns(2)
            for idx, item in enumerate(all_items):
                saved_val = saved_sec.get(item, custom.get(sec_name, {}).get(item, 0))
                with cols[idx % 2]:
                    v = st.number_input(item, min_value=0, value=int(saved_val or 0),
                                        step=500, key=f"be_{sec_name}_{item}", format="%d")
                sec_data[item] = v; sec_total += v

            st.markdown(f"**Sub-total: ${sec_total:,.0f}**")
            new_cfg["sections"][sec_name] = sec_data
            grand_total += sec_total
            if sec_name == "Technical and Sales Support":
                support_subtotal = sec_total

        # --- Global inputs ---
        st.markdown('<div class="sec-head">🔢 Program Parameters</div>', unsafe_allow_html=True)
        pc1, pc2, pc3 = st.columns(3)
        existing_partners = len(_load_partners())
        with pc1:
            num_p = st.number_input("Number of partners", min_value=1,
                value=int(cfg.get("num_partners") or existing_partners or 60), step=1, key="be_num_partners")
        with pc2:
            sup_calls = st.number_input("# of support calls (annual)", min_value=0,
                value=int(cfg.get("support_calls", 0)), step=100, key="be_sup_calls")
        with pc3:
            avg_min = st.number_input("Avg minutes per call", min_value=1,
                value=int(cfg.get("avg_min_per_call", 20)), step=1, key="be_avg_min")

        new_cfg["num_partners"] = num_p
        new_cfg["support_calls"] = sup_calls
        new_cfg["avg_min_per_call"] = avg_min

        st.markdown("---")
        _, bc = st.columns([3, 1])
        with bc: be_sub = st.form_submit_button("💾 Save Configuration", use_container_width=True, type="primary")

    if be_sub:
        _save_be(new_cfg); cfg = new_cfg
        st.session_state["_be_saved"] = True; st.rerun()

    # --- Results dashboard ---
    st.markdown("---")
    st.markdown("### 📊 Program Cost Summary")

    # Recalculate from cfg
    gt = sum(sum(v for v in items.values()) for items in cfg.get("sections", {}).values())
    np_ = cfg.get("num_partners", 1) or 1
    be_point = gt / np_
    sc_ = cfg.get("support_calls", 0) or 0
    am_ = cfg.get("avg_min_per_call", 20) or 20
    sup_cost = sum(cfg.get("sections", {}).get("Technical and Sales Support", {}).values())
    cpc = sup_cost / sc_ if sc_ > 0 else 0
    cpm = sup_cost / (sc_ * am_) if sc_ > 0 and am_ > 0 else 0

    m1, m2, m3 = st.columns(3)
    with m1: st.markdown(f'<div class="sum-card"><div class="sum-big">${gt:,.0f}</div><div class="sum-lbl">Total Program Costs</div></div>', unsafe_allow_html=True)
    with m2: st.markdown(f'<div class="sum-card"><div class="sum-big">{np_}</div><div class="sum-lbl">Number of Partners</div></div>', unsafe_allow_html=True)
    with m3: st.markdown(f'<div class="sum-card"><div class="sum-big" style="color:#49a34f">${be_point:,.2f}</div><div class="sum-lbl">Break-even per Partner</div></div>', unsafe_allow_html=True)

    if sup_cost > 0:
        st.markdown("#### 🛠️ Support Cost Metrics")
        s1, s2, s3, s4 = st.columns(4)
        with s1: st.metric("Support Costs", f"${sup_cost:,.0f}")
        with s2: st.metric("# Support Calls", f"{sc_:,}")
        with s3: st.metric("Cost per Call", f"${cpc:,.2f}")
        with s4: st.metric("Cost per Minute", f"${cpm:,.4f}")

    # Section breakdown table
    st.markdown("#### Cost Breakdown by Section")
    breakdown_rows = ""
    cfg_sections = list(cfg.get("sections", {}).keys())
    if not cfg_sections:
        cfg_sections = [s["section"] for s in BE_SECTIONS]
    for sec_name in cfg_sections:
        sec_items = cfg.get("sections", {}).get(sec_name, {})
        sec_t = sum(sec_items.values())
        pct = (sec_t / gt * 100) if gt > 0 else 0
        icon = BE_SECTION_ICONS.get(sec_name, "📁")
        breakdown_rows += f'<tr><td style="text-align:left;padding-left:10px">{icon} {sec_name}</td><td>${sec_t:,.0f}</td><td>{pct:.1f}%</td></tr>'
    if gt > 0:
        breakdown_rows += f'<tr class="hm-total"><td style="text-align:left;padding-left:10px;font-weight:800">TOTAL</td><td style="font-weight:800">${gt:,.0f}</td><td style="font-weight:800">100%</td></tr>'
    st.markdown(f'<table class="hm-tbl"><thead><tr><th style="text-align:left">Section</th><th>Cost</th><th>% of Total</th></tr></thead><tbody>{breakdown_rows}</tbody></table>', unsafe_allow_html=True)


# ═════════════════════════════════════════════════════════════════════════
# BREAK-EVEN — DETAILED ANALYSIS
# ═════════════════════════════════════════════════════════════════════════
elif page=="Break-even — Detailed Analysis":
    _brand(); st.markdown("## Break-even — Detailed Partner Cost Analysis")

    # Load break-even config for cost metrics
    be_cfg = _load_be()
    sup_cost = sum(be_cfg.get("sections", {}).get("Technical and Sales Support", {}).values())
    sc_ = be_cfg.get("support_calls", 0) or 0
    am_ = be_cfg.get("avg_min_per_call", 20) or 20
    cpm = sup_cost / (sc_ * am_) if sc_ > 0 and am_ > 0 else 0
    cpc = sup_cost / sc_ if sc_ > 0 else 0

    if sup_cost == 0:
        st.warning("⚠️ Complete **Break-even — Program Costs** first to set support cost metrics.")

    st.markdown("""<div class="info-box">
    Upload a CSV with columns: <b>Partner</b>, <b>Revenues</b>, <b># of calls</b>.
    Optional: <b>Time spent</b> (minutes). If missing, it will be estimated using your configured average minutes per call.
    Support cost per partner is calculated using cost-per-minute from your Program Costs configuration.</div>""", unsafe_allow_html=True)

    # Show current cost metrics
    cm1, cm2, cm3 = st.columns(3)
    with cm1: st.metric("Cost per Minute", f"${cpm:,.4f}" if cpm > 0 else "Not set")
    with cm2: st.metric("Cost per Call", f"${cpc:,.2f}" if cpc > 0 else "Not set")
    with cm3: st.metric("Avg Min/Call", f"{am_}")

    st.markdown("---")

    # Configurable avg minutes per call for this analysis
    avg_min_override = st.number_input("Average minutes per call (for estimating missing 'Time spent')",
        min_value=1, value=am_, step=1, key="da_avg_min")

    # File upload
    uploaded = st.file_uploader("📁 Upload Partner Cost CSV", type=["csv"], key="da_upload")

    # Try loading previously saved data
    df = None
    sd_path = _sd_path()

    if uploaded is not None:
        try:
            df = pd.read_csv(uploaded)
        except Exception as e:
            st.error(f"Error reading CSV: {e}"); df = None
    elif sd_path.exists():
        try:
            df = pd.read_csv(sd_path)
            st.info("📂 Loaded previously saved analysis data.")
        except: df = None

    if df is None:
        st.info("Upload a CSV to begin analysis, or complete one on the Program Costs page first.")
        st.stop()

    # Validate required columns
    col_map = {}
    for col in df.columns:
        cl = col.strip().lower()
        if cl in ("partner", "partner name"): col_map["Partner"] = col
        elif cl in ("revenues", "revenue"): col_map["Revenues"] = col
        elif cl in ("# of calls", "calls", "number of calls", "#calls"): col_map["Calls"] = col
        elif cl in ("time spent", "time", "minutes", "time spent (min)"): col_map["Time"] = col

    missing = [c for c in ["Partner", "Revenues", "Calls"] if c not in col_map]
    if missing:
        st.error(f"Missing required columns: {', '.join(missing)}. Found: {list(df.columns)}")
        st.stop()

    # Normalize column names
    df = df.rename(columns={col_map["Partner"]: "Partner", col_map["Revenues"]: "Revenues", col_map["Calls"]: "# of calls"})
    if "Time" in col_map:
        df = df.rename(columns={col_map["Time"]: "Time spent"})

    # Clean numeric columns
    for c in ["Revenues", "# of calls"]:
        df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0)
    if "Time spent" in df.columns:
        df["Time spent"] = pd.to_numeric(df["Time spent"], errors="coerce").fillna(0)
    else:
        df["Time spent"] = df["# of calls"] * avg_min_override

    # Fill missing time using avg
    df.loc[df["Time spent"] == 0, "Time spent"] = df.loc[df["Time spent"] == 0, "# of calls"] * avg_min_override

    # Sort by revenues descending
    df = df.sort_values("Revenues", ascending=False).reset_index(drop=True)

    # Calculate percentages and costs
    total_rev = df["Revenues"].sum()
    total_calls = df["# of calls"].sum()
    total_time = df["Time spent"].sum()

    df["% of revenues"] = (df["Revenues"] / total_rev * 100) if total_rev > 0 else 0
    df["% of calls"] = (df["# of calls"] / total_calls * 100) if total_calls > 0 else 0
    df["% of support time"] = (df["Time spent"] / total_time * 100) if total_time > 0 else 0

    if cpm > 0:
        df["Support cost"] = df["Time spent"] * cpm
    elif cpc > 0:
        df["Support cost"] = df["# of calls"] * cpc
    else:
        df["Support cost"] = 0

    total_support_cost = df["Support cost"].sum()
    df["% of cost"] = (df["Support cost"] / total_support_cost * 100) if total_support_cost > 0 else 0

    # Save processed data
    df.to_csv(sd_path, index=False)

    # --- Display results ---
    st.markdown("### 📊 Analysis Results")
    rm1, rm2, rm3, rm4 = st.columns(4)
    with rm1: st.metric("Partners", f"{len(df)}")
    with rm2: st.metric("Total Revenue", f"${total_rev:,.0f}")
    with rm3: st.metric("Total Calls", f"{int(total_calls):,}")
    with rm4: st.metric("Total Support Cost", f"${total_support_cost:,.2f}")

    # Display table
    st.markdown("### Partner Cost Table")
    display_df = df[["Partner", "Revenues", "% of revenues", "# of calls", "% of calls",
                     "Time spent", "% of support time", "Support cost", "% of cost"]].copy()

    # Add totals row
    totals = pd.DataFrame([{
        "Partner": "TOTAL", "Revenues": total_rev,
        "% of revenues": 100.0, "# of calls": total_calls,
        "% of calls": 100.0, "Time spent": total_time,
        "% of support time": 100.0, "Support cost": total_support_cost,
        "% of cost": 100.0
    }])
    display_with_totals = pd.concat([display_df, totals], ignore_index=True)

    # Format for display
    fmt_df = display_with_totals.copy()
    fmt_df["Revenues"] = fmt_df["Revenues"].apply(lambda x: f"${x:,.0f}")
    fmt_df["% of revenues"] = fmt_df["% of revenues"].apply(lambda x: f"{x:.1f}%")
    fmt_df["# of calls"] = fmt_df["# of calls"].apply(lambda x: f"{int(x):,}")
    fmt_df["% of calls"] = fmt_df["% of calls"].apply(lambda x: f"{x:.1f}%")
    fmt_df["Time spent"] = fmt_df["Time spent"].apply(lambda x: f"{int(x):,}")
    fmt_df["% of support time"] = fmt_df["% of support time"].apply(lambda x: f"{x:.1f}%")
    fmt_df["Support cost"] = fmt_df["Support cost"].apply(lambda x: f"${x:,.2f}")
    fmt_df["% of cost"] = fmt_df["% of cost"].apply(lambda x: f"{x:.1f}%")

    st.dataframe(fmt_df, use_container_width=True, hide_index=True)

    # Downloads
    dl1, dl2 = st.columns(2)
    with dl1:
        csv_buf = display_with_totals.to_csv(index=False)
        st.download_button("⬇️ Download CSV", csv_buf, "partner_cost_analysis.csv", "text/csv", type="primary")
    with dl2:
        try:
            import openpyxl
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Partner Costs"
            hf = PatternFill(start_color="1E2A3A", end_color="1E2A3A", fill_type="solid")
            hfont = Font(color="FFFFFF", bold=True, size=10)
            bdr = Border(left=Side(style="thin", color="CCCCCC"), right=Side(style="thin", color="CCCCCC"),
                         top=Side(style="thin", color="CCCCCC"), bottom=Side(style="thin", color="CCCCCC"))
            headers = list(display_with_totals.columns)
            for ci, h in enumerate(headers, 1):
                c = ws.cell(1, ci, h); c.fill = hf; c.font = hfont; c.border = bdr
                c.alignment = Alignment(horizontal="center", wrap_text=True)
                ws.column_dimensions[c.column_letter].width = 16
            ws.column_dimensions["A"].width = 24
            for ri, (_, row) in enumerate(display_with_totals.iterrows(), 2):
                for ci, h in enumerate(headers, 1):
                    v = row[h]
                    cell = ws.cell(ri, ci, v); cell.border = bdr
                    cell.alignment = Alignment(horizontal="center")
                    if h in ("Revenues", "Support cost") and isinstance(v, (int, float)):
                        cell.number_format = '#,##0'
                    elif "%" in h and isinstance(v, (int, float)):
                        cell.value = v / 100; cell.number_format = "0.0%"
                    if row["Partner"] == "TOTAL":
                        cell.font = Font(bold=True)
            buf = io.BytesIO(); wb.save(buf)
            st.download_button("⬇️ Download Excel", buf.getvalue(), "partner_cost_analysis.xlsx",
                              "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        except ImportError: pass

    # --- Visualizations ---
    st.markdown("---")
    st.markdown("### 📈 Visualizations")

    chart_df = df.head(15).copy()
    if len(chart_df) > 0 and total_support_cost > 0:
        try:
            import altair as alt

            tab1, tab2, tab3 = st.tabs(["Support Cost vs Revenue", "Cost Distribution", "Cost/Revenue Ratio"])

            with tab1:
                st.markdown("#### Top Partners: Support Cost vs Revenue")
                # Melt for grouped bar chart
                bar_src = chart_df[["Partner", "Revenues", "Support cost"]].copy()
                bar_src = bar_src.melt(id_vars="Partner", var_name="Metric", value_name="Amount")
                bar_chart = alt.Chart(bar_src).mark_bar().encode(
                    x=alt.X("Partner:N", sort=list(chart_df["Partner"]), axis=alt.Axis(labelAngle=-45, labelLimit=120)),
                    y=alt.Y("Amount:Q", title="$"),
                    color=alt.Color("Metric:N", scale=alt.Scale(domain=["Revenues","Support cost"], range=["#2563eb","#dc4040"])),
                    xOffset="Metric:N",
                    tooltip=["Partner", "Metric", alt.Tooltip("Amount:Q", format="$,.0f")]
                ).properties(height=400)
                st.altair_chart(bar_chart, use_container_width=True)

            with tab2:
                st.markdown("#### Cost Distribution by Partner")
                top10 = df.head(10).copy()
                rest_cost = df.iloc[10:]["Support cost"].sum() if len(df) > 10 else 0
                pie_src = top10[["Partner", "Support cost"]].copy()
                if rest_cost > 0:
                    pie_src = pd.concat([pie_src, pd.DataFrame([{"Partner": "Others", "Support cost": rest_cost}])], ignore_index=True)
                pie_chart = alt.Chart(pie_src).mark_arc(innerRadius=50).encode(
                    theta=alt.Theta("Support cost:Q"),
                    color=alt.Color("Partner:N", legend=alt.Legend(title="Partner")),
                    tooltip=["Partner", alt.Tooltip("Support cost:Q", format="$,.2f")]
                ).properties(height=400)
                st.altair_chart(pie_chart, use_container_width=True)

                # Revenue distribution pie
                st.markdown("#### Revenue Distribution by Partner")
                rev_src = top10[["Partner", "Revenues"]].copy()
                rest_rev = df.iloc[10:]["Revenues"].sum() if len(df) > 10 else 0
                if rest_rev > 0:
                    rev_src = pd.concat([rev_src, pd.DataFrame([{"Partner": "Others", "Revenues": rest_rev}])], ignore_index=True)
                rev_pie = alt.Chart(rev_src).mark_arc(innerRadius=50).encode(
                    theta=alt.Theta("Revenues:Q"),
                    color=alt.Color("Partner:N", legend=alt.Legend(title="Partner")),
                    tooltip=["Partner", alt.Tooltip("Revenues:Q", format="$,.0f")]
                ).properties(height=400)
                st.altair_chart(rev_pie, use_container_width=True)

            with tab3:
                st.markdown("#### Cost / Revenue Ratio by Partner")
                ratio_df = chart_df[["Partner", "Revenues", "# of calls", "Support cost"]].copy()
                ratio_df["Cost/Rev %"] = ratio_df.apply(lambda r: (r["Support cost"] / r["Revenues"] * 100) if r["Revenues"] > 0 else 0, axis=1)
                ratio_chart = alt.Chart(ratio_df).mark_bar().encode(
                    x=alt.X("Partner:N", sort=list(chart_df["Partner"]), axis=alt.Axis(labelAngle=-45, labelLimit=120)),
                    y=alt.Y("Cost/Rev %:Q", title="Support Cost as % of Revenue"),
                    color=alt.condition(
                        alt.datum["Cost/Rev %"] > 10, alt.value("#dc4040"),
                        alt.condition(alt.datum["Cost/Rev %"] > 5, alt.value("#d4a917"), alt.value("#1b6e23"))
                    ),
                    tooltip=["Partner", alt.Tooltip("Revenues:Q", format="$,.0f"),
                             alt.Tooltip("Support cost:Q", format="$,.2f"),
                             alt.Tooltip("Cost/Rev %:Q", format=".1f")]
                ).properties(height=400)
                st.altair_chart(ratio_chart, use_container_width=True)

                # Table view
                st.markdown("##### Detail")
                tbl_html = '<table class="hm-tbl"><thead><tr><th style="text-align:left">Partner</th><th>Revenue</th><th>Calls</th><th>Support Cost</th><th>Cost/Rev %</th></tr></thead><tbody>'
                for _, r in ratio_df.iterrows():
                    rc = "#dc4040" if r["Cost/Rev %"] > 10 else "#d4a917" if r["Cost/Rev %"] > 5 else "#1b6e23"
                    tbl_html += f'<tr><td style="text-align:left;padding-left:10px">{r["Partner"]}</td><td>${r["Revenues"]:,.0f}</td><td>{int(r["# of calls"]):,}</td><td>${r["Support cost"]:,.2f}</td><td style="color:{rc};font-weight:700">{r["Cost/Rev %"]:.1f}%</td></tr>'
                tbl_html += '</tbody></table>'
                st.markdown(tbl_html, unsafe_allow_html=True)

        except Exception:
            # Altair unavailable — fall back to native Streamlit charts
            st.markdown("#### Top Partners: Support Cost vs Revenue")
            fb_df = chart_df[["Partner", "Revenues", "Support cost"]].set_index("Partner")
            st.bar_chart(fb_df)

            st.markdown("#### Cost / Revenue Ratio")
            ratio_df = chart_df[["Partner", "Revenues", "# of calls", "Support cost"]].copy()
            ratio_df["Cost/Rev %"] = ratio_df.apply(lambda r: (r["Support cost"] / r["Revenues"] * 100) if r["Revenues"] > 0 else 0, axis=1)
            st.bar_chart(ratio_df[["Partner","Cost/Rev %"]].set_index("Partner"))

            st.markdown("##### Detail")
            tbl_html = '<table class="hm-tbl"><thead><tr><th style="text-align:left">Partner</th><th>Revenue</th><th>Calls</th><th>Support Cost</th><th>Cost/Rev %</th></tr></thead><tbody>'
            for _, r in ratio_df.iterrows():
                rc = "#dc4040" if r["Cost/Rev %"] > 10 else "#d4a917" if r["Cost/Rev %"] > 5 else "#1b6e23"
                tbl_html += f'<tr><td style="text-align:left;padding-left:10px">{r["Partner"]}</td><td>${r["Revenues"]:,.0f}</td><td>{int(r["# of calls"]):,}</td><td>${r["Support cost"]:,.2f}</td><td style="color:{rc};font-weight:700">{r["Cost/Rev %"]:.1f}%</td></tr>'
            tbl_html += '</tbody></table>'
            st.markdown(tbl_html, unsafe_allow_html=True)
    else:
        st.info("Add support cost data in Program Costs and upload partner data to see visualizations.")


# ═════════════════════════════════════════════════════════════════════════
# REVENUE RECOVERY — margin recapture from non-performing longtail
# ═════════════════════════════════════════════════════════════════════════
elif page == "Revenue Recovery":
    _brand()
    st.markdown("## Strategic Margin Realignment & Revenue Recovery")
    st.markdown("""<div class="info-box">
    Identify partners who have been <b>grandfathered into high margins</b> despite low performance
    on net-new logo generation. Calculate the potential savings if their margin is reset to a
    baseline rate. Only partners with a <b>Partner Discount</b> (margin %) and
    <b>Annual Revenue</b> on file are analyzed.</div>""", unsafe_allow_html=True)

    # Gate: need partner data
    if not _raw_path().exists():
        st.warning("No partner data found. Import partners first via the **Import Data** page.")
        st.stop()

    raw_all = _load_raw()
    if not raw_all:
        st.warning("No partner data found. Import partners first via the **Import Data** page.")
        st.stop()

    # ── Controls ──
    st.markdown("---")
    st.markdown("### Parameters")
    ctrl1, ctrl2 = st.columns(2)
    with ctrl1:
        nn_threshold = st.number_input(
            "Net-New Logo Revenue threshold ($)",
            min_value=0, max_value=10_000_000, value=50_000, step=5_000,
            help="Partners with net-new logo revenue **below** this amount are flagged as non-performers.",
            key="rr_nn_threshold",
        )
    with ctrl2:
        baseline_margin = st.number_input(
            "Proposed baseline margin (%)",
            min_value=0.0, max_value=100.0, value=10.0, step=1.0,
            help="The new margin rate to apply to non-performing partners.",
            key="rr_baseline",
        )

    # ── Calculation ──
    df = _calculate_revenue_recovery(raw_all, float(nn_threshold), float(baseline_margin))

    if df.empty:
        st.info("No non-performing partners found with the current threshold. "
                "Try lowering the net-new logo revenue threshold or check that partners have "
                "a **Partner Discount** and **Annual Revenue** on file.")
        st.stop()

    total_recapture = df["Recapture $"].sum()
    total_current = df["Current Margin $"].sum()
    total_proposed = df["New Margin $"].sum()
    partner_count = len(df)

    # ── Summary cards ──
    st.markdown("---")
    st.markdown("### Summary")
    k1, k2, k3, k4 = st.columns(4)
    with k1:
        st.metric("Non-Performing Partners", f"{partner_count:,}")
    with k2:
        st.metric("Current Margin Cost", f"${total_current:,.0f}")
    with k3:
        st.metric("Proposed Margin Cost", f"${total_proposed:,.0f}")
    with k4:
        st.markdown(
            f'<div style="background:#1b6e23;color:#fff;border-radius:10px;padding:16px 20px;text-align:center;">'
            f'<div style="font-size:.85rem;opacity:.85;">Total Potential Recapture</div>'
            f'<div style="font-size:1.8rem;font-weight:700;margin-top:4px;">${total_recapture:,.0f}</div></div>',
            unsafe_allow_html=True,
        )

    # ── Detail table ──
    st.markdown("---")
    st.markdown("### Partner Detail")

    # Format for display
    display_df = df.copy()
    display_df.index = range(1, len(display_df) + 1)
    display_df.index.name = "#"

    if HAS_AGGRID:
        gb = GridOptionsBuilder.from_dataframe(display_df)
        gb.configure_default_column(resizable=True, sortable=True, filterable=True)
        gb.configure_column("Partner", pinned="left", width=200)
        for col in ["Annual Revenue", "Current Margin $", "New Margin $", "Recapture $", "Net-New Logo Revenue"]:
            gb.configure_column(col, type=["numericColumn"], valueFormatter="'$' + Math.round(value).toLocaleString()")
        for col in ["Current Margin %", "New Margin %"]:
            gb.configure_column(col, type=["numericColumn"], valueFormatter="value.toFixed(1) + '%'")
        gb.configure_column("Recapture $", cellStyle=JsCode("""
            function(params) {
                if (params.value > 0) return {'color': '#1b6e23', 'fontWeight': '700'};
                return {};
            }
        """))
        grid_opts = gb.build()
        grid_opts["domLayout"] = "autoHeight"
        AgGrid(display_df, gridOptions=grid_opts, fit_columns_on_grid_load=True,
               update_mode=GridUpdateMode.NO_UPDATE, columns_auto_size_mode=ColumnsAutoSizeMode.FIT_CONTENTS,
               allow_unsafe_jscode=True)
    else:
        # Fallback: HTML table
        tbl = '<table class="hm-tbl"><thead><tr>'
        tbl += '<th style="text-align:left">Partner</th>'
        tbl += '<th>Annual Revenue</th><th>Current %</th><th>Current $</th>'
        tbl += '<th>New %</th><th>New $</th><th>Recapture $</th><th>Net-New Logo $</th></tr></thead><tbody>'
        for _, r in display_df.iterrows():
            recap_color = "#1b6e23" if r["Recapture $"] > 0 else "#dc4040"
            tbl += (
                f'<tr>'
                f'<td style="text-align:left;padding-left:10px">{r["Partner"]}</td>'
                f'<td>${r["Annual Revenue"]:,.0f}</td>'
                f'<td>{r["Current Margin %"]:.1f}%</td>'
                f'<td>${r["Current Margin $"]:,.0f}</td>'
                f'<td>{r["New Margin %"]:.1f}%</td>'
                f'<td>${r["New Margin $"]:,.0f}</td>'
                f'<td style="color:{recap_color};font-weight:700">${r["Recapture $"]:,.0f}</td>'
                f'<td>${r["Net-New Logo Revenue"]:,.0f}</td>'
                f'</tr>'
            )
        tbl += '</tbody></table>'
        st.markdown(tbl, unsafe_allow_html=True)

    # ── Total recapture bar (repeat at bottom for prominence) ──
    st.markdown(
        f'<div style="background:linear-gradient(135deg,#1b6e23,#2e8b57);color:#fff;border-radius:10px;'
        f'padding:20px 28px;margin-top:16px;display:flex;justify-content:space-between;align-items:center;">'
        f'<div><span style="font-size:1.1rem;font-weight:600;">Total Potential Revenue Recapture</span><br>'
        f'<span style="opacity:.8;font-size:.88rem;">{partner_count} non-performing partner(s) '
        f'&bull; baseline margin reset to {baseline_margin:.0f}%</span></div>'
        f'<div style="font-size:2rem;font-weight:800;">${total_recapture:,.0f}</div></div>',
        unsafe_allow_html=True,
    )

    # ── Export ──
    st.markdown("---")
    if is_admin:
        csv_data = df.to_csv(index=False)
        st.download_button(
            "Download Report (CSV)",
            csv_data,
            file_name="revenue_recovery_report.csv",
            mime="text/csv",
            use_container_width=False,
        )
    else:
        st.info("🔒 Exporting is disabled in the trial environment. "
                "Contact your York Group consultant to enable data exports.")

# ═════════════════════════════════════════════════════════════════════════
# USER GUIDE
# ═════════════════════════════════════════════════════════════════════════
elif page == "User Guide":
    _brand()
    from pathlib import Path
    _guide_path = Path(__file__).parent / "USER_GUIDE.md"
    if _guide_path.exists():
        st.markdown(_guide_path.read_text(), unsafe_allow_html=False)
    else:
        st.warning("User guide file not found.")
